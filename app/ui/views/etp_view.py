"""
etp_view.py — Onglet Tableau de Charge (ETP)
Affiche la charge de travail par semaine/mois sur projets et tâches.
Heures + conversion jours (base 7h/jour).
Export Excel via openpyxl.
"""
import logging
from datetime import datetime, date, timedelta
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QSpinBox,
    QGroupBox, QFormLayout, QHeaderView, QMessageBox,
    QAbstractItemView, QSizePolicy, QFrame
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from app.services.database_service import db_service

logger = logging.getLogger(__name__)

HEURES_PAR_JOUR = 7.0   # 1 ETP = 7h/jour
HEURES_PAR_MOIS = 154.0  # ~22 jours ouvrés × 7h
COULEURS = {
    'header':   '#2c3e50',
    'projet':   '#2980b9',
    'tache_bc': '#8e44ad',
    'tache':    '#27ae60',
    'total':    '#e67e22',
    'surcharge':'#e74c3c',
    'ok':       '#27ae60',
    'vide':     '#ecf0f1',
}

def _h2j(heures):
    """Convertit des heures en jours (arrondi 2 déc.)."""
    return round(heures / HEURES_PAR_JOUR, 2)

def _label_h_j(heures):
    """Formate : 14h = 14h (2.0j)."""
    if heures <= 0:
        return ''
    return f"{heures:.1f}h ({_h2j(heures):.1f}j)"

class ETPView(QWidget):
    """Vue tableau de charge ETP — heures + jours."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.annee   = date.today().year
        self.mode    = 'mensuel'
        try:
            self.setup_ui()
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            from PyQt5.QtWidgets import QVBoxLayout, QLabel
            lay = QVBoxLayout(self)
            lay.addWidget(QLabel(f"Erreur init ETP: {e}"))
            return
        try:
            self.load_data()
        except Exception as e:
            import traceback
            print(traceback.format_exc())

    # ── UI ────────────────────────────────────────────────────────────────────

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        titre = QLabel("Tableau de Charge ETP")
        titre.setStyleSheet("font-size:18px;font-weight:bold;color:#2c3e50;padding:6px;")
        layout.addWidget(titre)

        ctrl = QHBoxLayout()

        ctrl.addWidget(QLabel("Année :"))
        self.spin_annee = QSpinBox()
        self.spin_annee.setRange(2020, 2035)
        self.spin_annee.setValue(self.annee)
        self.spin_annee.valueChanged.connect(self._on_annee_change)
        ctrl.addWidget(self.spin_annee)

        ctrl.addWidget(QLabel("   Vue :"))
        self.combo_mode = QComboBox()
        self.combo_mode.addItems(["Mensuel", "Hebdomadaire"])
        self.combo_mode.currentIndexChanged.connect(self._on_mode_change)
        ctrl.addWidget(self.combo_mode)

        ctrl.addSpacing(20)
        btn_refresh = QPushButton("🔄 Actualiser")
        btn_refresh.setStyleSheet(
            "background:#3498db;color:white;font-weight:bold;padding:6px 14px;")
        btn_refresh.clicked.connect(self.load_data)
        ctrl.addWidget(btn_refresh)

        btn_export = QPushButton("📥 Export Excel")
        btn_export.setStyleSheet(
            "background:#27ae60;color:white;font-weight:bold;padding:6px 14px;")
        btn_export.clicked.connect(self.export_excel)
        ctrl.addWidget(btn_export)

        ctrl.addStretch()
        layout.addLayout(ctrl)

        # ── KPI ───────────────────────────────────────────────────────────────
        kpi_frame = QFrame()
        kpi_frame.setStyleSheet(
            "background:#f8f9fa;border:1px solid #dee2e6;border-radius:6px;padding:8px;")
        kpi_layout = QHBoxLayout(kpi_frame)

        self.kpi_projets  = self._kpi("📁 Projets actifs", "0")
        self.kpi_taches   = self._kpi("✅ Tâches en cours", "0")
        self.kpi_h_total  = self._kpi("⏱️ Total heures", "0h")
        self.kpi_j_total  = self._kpi("📅 Total jours", "0j")
        self.kpi_etp      = self._kpi("👤 ETP annuel", "0.0")

        for kpi in [self.kpi_projets, self.kpi_taches,
                    self.kpi_h_total, self.kpi_j_total, self.kpi_etp]:
            kpi_layout.addWidget(kpi)

        layout.addWidget(kpi_frame)

        # ── Tableau ───────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setDefaultSectionSize(90)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(
            "QTableWidget { gridline-color: #dee2e6; font-size: 12px; }"
            "QHeaderView::section { background-color: #2c3e50; color: white;"
            " font-weight: bold; padding: 6px; border: 1px solid #1a252f; }"
        )
        layout.addWidget(self.table)

        # ── Légende ───────────────────────────────────────────────────────────
        leg = QHBoxLayout()
        for couleur, texte in [
            ('#2980b9', '■ Projet'),
            ('#8e44ad', '■ Tâche liée à BC'),
            ('#27ae60', '■ Tâche standard'),
            ('#e67e22', '■ Total'),
            ('#e74c3c', '■ Surcharge (>100%)'),
        ]:
            lbl = QLabel(texte)
            lbl.setStyleSheet(f"color:{couleur};font-weight:bold;margin-right:12px;")
            leg.addWidget(lbl)
        leg.addStretch()
        layout.addLayout(leg)

    def _kpi(self, titre, valeur):
        """Crée un widget KPI."""
        w = QWidget()
        w.setStyleSheet(
            "background:white;border:1px solid #dee2e6;border-radius:6px;padding:8px;")
        v = QVBoxLayout(w)
        v.setContentsMargins(8, 4, 8, 4)
        lbl_titre = QLabel(titre)
        lbl_titre.setStyleSheet("font-size:11px;color:#6c757d;")
        lbl_val   = QLabel(valeur)
        lbl_val.setStyleSheet("font-size:18px;font-weight:bold;color:#2c3e50;")
        lbl_val.setObjectName("val")
        v.addWidget(lbl_titre)
        v.addWidget(lbl_val)
        return w

    def _set_kpi(self, widget, valeur):
        widget.findChild(QLabel, "val").setText(str(valeur))

    # ── Contrôles ─────────────────────────────────────────────────────────────

    def _on_annee_change(self, val):
        self.annee = val
        self.load_data()

    def _on_mode_change(self, idx):
        self.mode = 'hebdo' if idx == 1 else 'mensuel'
        self.load_data()

    # ── Chargement données ────────────────────────────────────────────────────

    def load_data(self):
        """Charge et affiche le tableau de charge."""
        try:
            periodes = self._get_periodes()
            lignes   = self._get_lignes_charge()
            self._build_table(periodes, lignes)
            self._update_kpi(lignes)
        except Exception as e:
            logger.error(f"Erreur chargement ETP: {e}", exc_info=True)
            QMessageBox.warning(self, "Erreur", f"Impossible de charger les données ETP:\n{e}")

    def _get_periodes(self):
        """Retourne la liste des colonnes (mois ou semaines)."""
        if self.mode == 'mensuel':
            return [
                {'label': date(self.annee, m, 1).strftime('%b %Y'),
                 'debut': date(self.annee, m, 1),
                 'fin':   date(self.annee, m + 1, 1) - timedelta(days=1)
                          if m < 12 else date(self.annee, 12, 31)}
                for m in range(1, 13)
            ]
        else:
            # Semaines de l'année
            periodes = []
            d = date(self.annee, 1, 1)
            while d.isocalendar()[0] == self.annee or d.year == self.annee:
                lundi = d - timedelta(days=d.weekday())
                vendredi = lundi + timedelta(days=4)
                if lundi.year == self.annee or vendredi.year == self.annee:
                    periodes.append({
                        'label': f"S{lundi.isocalendar()[1]:02d}",
                        'debut': lundi,
                        'fin':   vendredi
                    })
                d += timedelta(weeks=1)
                if len(periodes) >= 53:
                    break
            return periodes

    def _get_lignes_charge(self):
        """
        Retourne les lignes de charge :
        - Projets actifs avec leurs heures estimées réparties
        - Tâches en cours (avec ou sans projet)
        Chaque ligne = {'type', 'label', 'sous_label', 'heures_par_periode': [...]}
        """
        lignes = []

        # ── Projets ────────────────────────────────────────────────────────
        try:
            projets = db_service.fetch_all("""
                SELECT p.id, p.nom, p.code,
                       COALESCE(p.statut, 'EN_COURS') AS statut,
                       p.date_debut, p.date_fin_prevue,
                       COALESCE(SUM(t.estimation_heures), 0) AS h_estimees,
                       COALESCE(SUM(t.heures_reelles),    0) AS h_reelles,
                       COUNT(t.id) AS nb_taches
                FROM projets p
                LEFT JOIN taches t ON t.projet_id = p.id
                WHERE COALESCE(p.statut,'') NOT IN ('TERMINE','ABANDONNE','ANNULE')
                GROUP BY p.id
                ORDER BY p.nom
            """) or []
        except Exception:
            projets = []

        for proj in projets:
            p = dict(proj)
            periodes = self._get_periodes()
            heures_pp = self._repartir_heures(
                float(p.get('h_estimees') or 0),
                p.get('date_debut'), p.get('date_fin_prevue'),
                periodes)
            lignes.append({
                'type':      'projet',
                'label':     f"📁 {p.get('nom','')}",
                'sous_label': p.get('statut',''),
                'heures':    float(p.get('h_estimees') or 0),
                'heures_pp': heures_pp,
                'projet_id': p['id'],
            })

        # ── Tâches ─────────────────────────────────────────────────────────
        try:
            taches = db_service.fetch_all("""
                SELECT t.id, t.titre, t.statut, t.projet_id,
                       t.date_debut, t.date_echeance,
                       t.estimation_heures, t.heures_reelles,
                       t.tags,
                       p.nom AS projet_nom
                FROM taches t
                LEFT JOIN projets p ON p.id = t.projet_id
                WHERE t.statut NOT IN ('TERMINE','ANNULE')
                ORDER BY CASE WHEN t.date_echeance IS NULL THEN 1 ELSE 0 END, t.date_echeance, t.titre
            """) or []
        except Exception:
            taches = []

        for tache in taches:
            t = dict(tache)
            est_bc = 'BC' in (t.get('tags') or '') or 'bc' in (t.get('titre') or '').lower()
            periodes = self._get_periodes()
            heures_pp = self._repartir_heures(
                float(t.get('estimation_heures') or 0),
                t.get('date_debut'), t.get('date_echeance'),
                periodes)
            lignes.append({
                'type':      'tache_bc' if est_bc else 'tache',
                'label':     f"  {'🛒' if est_bc else '✅'} {t.get('titre','')}",
                'sous_label': t.get('projet_nom') or '(sans projet)',
                'heures':    float(t.get('estimation_heures') or 0),
                'heures_pp': heures_pp,
            })

        return lignes

    def _repartir_heures(self, total_h, date_debut, date_fin, periodes):
        """Répartit les heures uniformément sur les périodes couvertes."""
        heures_pp = [0.0] * len(periodes)
        if not total_h:
            return heures_pp

        # Parser les dates
        def _parse(d):
            if not d:
                return None
            if isinstance(d, date):
                return d
            try:
                return datetime.strptime(str(d)[:10], '%Y-%m-%d').date()
            except Exception:
                return None

        d_debut = _parse(date_debut) or date(self.annee, 1, 1)
        d_fin   = _parse(date_fin)   or date(self.annee, 12, 31)

        # Calculer les jours de chevauchement par période
        jours_par_periode = []
        for p in periodes:
            debut_p = max(d_debut, p['debut'])
            fin_p   = min(d_fin,   p['fin'])
            jours   = max(0, (fin_p - debut_p).days + 1)
            jours_par_periode.append(jours)

        total_jours = sum(jours_par_periode)
        if total_jours <= 0:
            # Projet hors période → colonne la plus proche
            heures_pp[0] = total_h
            return heures_pp

        for i, j in enumerate(jours_par_periode):
            heures_pp[i] = round(total_h * j / total_jours, 1)

        return heures_pp

    # ── Construction table ────────────────────────────────────────────────────

    def _build_table(self, periodes, lignes):
        """Construit le QTableWidget."""
        cols = ['Projet / Tâche', 'Statut', 'Total'] + [p['label'] for p in periodes]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setColumnWidth(1, 90)
        self.table.setColumnWidth(2, 110)

        # Calculer totaux par période
        totaux = [0.0] * len(periodes)
        rows_data = []
        for ligne in lignes:
            rows_data.append(ligne)
            for i, h in enumerate(ligne['heures_pp']):
                totaux[i] += h

        # +1 ligne total
        self.table.setRowCount(len(rows_data) + 1)

        for row_idx, ligne in enumerate(rows_data):
            couleur = QColor(COULEURS.get(ligne['type'], '#ffffff'))
            couleur_texte = QColor('#ffffff')

            # Colonne 0 : label
            item = QTableWidgetItem(ligne['label'])
            item.setBackground(couleur)
            item.setForeground(couleur_texte)
            item.setFont(QFont("Arial", 10,
                               QFont.Bold if ligne['type'] == 'projet' else QFont.Normal))
            self.table.setItem(row_idx, 0, item)

            # Colonne 1 : sous-label statut/projet
            item2 = QTableWidgetItem(ligne['sous_label'])
            item2.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 1, item2)

            # Colonne 2 : total heures
            item3 = QTableWidgetItem(_label_h_j(ligne['heures']))
            item3.setTextAlignment(Qt.AlignCenter)
            item3.setFont(QFont("Arial", 10, QFont.Bold))
            self.table.setItem(row_idx, 2, item3)

            # Colonnes périodes
            for col_idx, h in enumerate(ligne['heures_pp']):
                item_p = QTableWidgetItem(_label_h_j(h) if h > 0 else '')
                item_p.setTextAlignment(Qt.AlignCenter)
                if h > 0:
                    # Fond légèrement coloré selon type
                    c = QColor(couleur)
                    c.setAlpha(40)
                    item_p.setBackground(c)
                self.table.setItem(row_idx, 3 + col_idx, item_p)

        # Ligne total
        row_total = len(rows_data)
        total_h = sum(totaux)

        item_t = QTableWidgetItem("📊 TOTAL")
        item_t.setBackground(QColor(COULEURS['total']))
        item_t.setForeground(QColor('#ffffff'))
        item_t.setFont(QFont("Arial", 11, QFont.Bold))
        self.table.setItem(row_total, 0, item_t)

        item_t2 = QTableWidgetItem("")
        item_t2.setBackground(QColor(COULEURS['total']))
        self.table.setItem(row_total, 1, item_t2)

        item_t3 = QTableWidgetItem(_label_h_j(total_h))
        item_t3.setTextAlignment(Qt.AlignCenter)
        item_t3.setBackground(QColor(COULEURS['total']))
        item_t3.setForeground(QColor('#ffffff'))
        item_t3.setFont(QFont("Arial", 11, QFont.Bold))
        self.table.setItem(row_total, 2, item_t3)

        for col_idx, h in enumerate(totaux):
            # Surcharge > 100% d'un ETP mensuel
            limite = HEURES_PAR_MOIS if self.mode == 'mensuel' else 35.0
            couleur_t = COULEURS['surcharge'] if h > limite else COULEURS['total']
            item_p = QTableWidgetItem(_label_h_j(h) if h > 0 else '')
            item_p.setTextAlignment(Qt.AlignCenter)
            item_p.setBackground(QColor(couleur_t))
            item_p.setForeground(QColor('#ffffff'))
            item_p.setFont(QFont("Arial", 10, QFont.Bold))
            self.table.setItem(row_total, 3 + col_idx, item_p)

        self.table.resizeRowsToContents()

    def _update_kpi(self, lignes):
        """Met à jour les KPI."""
        nb_projets = sum(1 for l in lignes if l['type'] == 'projet')
        nb_taches  = sum(1 for l in lignes if l['type'] in ('tache','tache_bc'))
        total_h    = sum(l['heures'] for l in lignes if l['type'] in ('tache','tache_bc'))
        etp        = round(total_h / HEURES_PAR_MOIS / 12, 2) if total_h else 0

        self._set_kpi(self.kpi_projets, str(nb_projets))
        self._set_kpi(self.kpi_taches,  str(nb_taches))
        self._set_kpi(self.kpi_h_total, f"{total_h:.0f}h")
        self._set_kpi(self.kpi_j_total, f"{_h2j(total_h):.1f}j")
        self._set_kpi(self.kpi_etp,     f"{etp:.2f}")

    # ── Export Excel ──────────────────────────────────────────────────────────

    def export_excel(self):
        """Exporte le tableau de charge en Excel."""
        try:
            import openpyxl
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter
            from PyQt5.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.warning(self, "Module manquant",
                "openpyxl n'est pas installé.\npip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter tableau de charge",
            f"tableau_charge_ETP_{self.annee}.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return

        try:
            periodes = self._get_periodes()
            lignes   = self._get_lignes_charge()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Charge ETP {self.annee}"

            # Styles
            def fill(hex_color):
                return PatternFill("solid", fgColor=hex_color.lstrip('#'))

            font_bold  = Font(bold=True, color="FFFFFF", size=11)
            font_titre = Font(bold=True, color="FFFFFF", size=13)
            font_norm  = Font(size=10)
            align_c    = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
            align_l    = Alignment(horizontal="left",   vertical="center",
                                   wrap_text=True)
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            colors = {
                'projet':   '2980B9',
                'tache_bc': '8E44AD',
                'tache':    '27AE60',
                'total':    'E67E22',
                'header':   '2C3E50',
                'surcharge':'E74C3C',
            }

            # ── Titre ────────────────────────────────────────────────────────
            ws.merge_cells(f"A1:{get_column_letter(3 + len(periodes))}1")
            c = ws.cell(1, 1,
                f"Tableau de Charge ETP — {self.annee} "
                f"({'Mensuel' if self.mode == 'mensuel' else 'Hebdomadaire'})")
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.fill = fill(colors['header'])
            c.alignment = align_c
            ws.row_dimensions[1].height = 28

            # ── En-têtes colonnes ─────────────────────────────────────────────
            headers = ['Projet / Tâche', 'Statut', 'Total (h+j)'] + \
                      [p['label'] for p in periodes]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(2, col_idx, h)
                c.font = font_bold
                c.fill = fill(colors['header'])
                c.alignment = align_c
                c.border    = border
            ws.row_dimensions[2].height = 22

            # ── Données ───────────────────────────────────────────────────────
            totaux = [0.0] * len(periodes)
            for row_idx, ligne in enumerate(lignes, 3):
                typ  = ligne['type']
                col  = colors.get(typ, 'FFFFFF')
                font = Font(bold=(typ=='projet'), color='FFFFFF', size=10)

                c = ws.cell(row_idx, 1, ligne['label'])
                c.font = font; c.fill = fill(col)
                c.alignment = align_l; c.border = border

                c = ws.cell(row_idx, 2, ligne['sous_label'])
                c.font = font; c.fill = fill(col)
                c.alignment = align_c; c.border = border

                total_txt = _label_h_j(ligne['heures'])
                c = ws.cell(row_idx, 3, total_txt)
                c.font = Font(bold=True, color='FFFFFF', size=10)
                c.fill = fill(col); c.alignment = align_c; c.border = border

                for col_idx, h in enumerate(ligne['heures_pp']):
                    totaux[col_idx] += h
                    txt = _label_h_j(h) if h > 0 else ''
                    c = ws.cell(row_idx, 4 + col_idx, txt)
                    c.font = Font(size=10, bold=False)
                    c.alignment = align_c; c.border = border
                    if h > 0:
                        # Fond léger
                        r, g, b = (
                            int(col[0:2],16),
                            int(col[2:4],16),
                            int(col[4:6],16))
                        light = f"{min(255,r+150):02X}{min(255,g+150):02X}{min(255,b+150):02X}"
                        c.fill = fill(light)

                ws.row_dimensions[row_idx].height = 18

            # ── Ligne total ───────────────────────────────────────────────────
            row_t = len(lignes) + 3
            total_h = sum(totaux)
            limite  = HEURES_PAR_MOIS if self.mode == 'mensuel' else 35.0

            for col_idx in range(1, 4 + len(periodes)):
                c = ws.cell(row_t, col_idx)
                c.font = Font(bold=True, color='FFFFFF', size=11)
                c.fill = fill(colors['total'])
                c.alignment = align_c; c.border = border

            ws.cell(row_t, 1, "📊 TOTAL")
            ws.cell(row_t, 3, _label_h_j(total_h))
            for col_idx, h in enumerate(totaux):
                col_sur = colors['surcharge'] if h > limite else colors['total']
                c = ws.cell(row_t, 4 + col_idx, _label_h_j(h) if h > 0 else '')
                c.fill = fill(col_sur)
                c.font = Font(bold=True, color='FFFFFF', size=10)
                c.alignment = align_c; c.border = border

            ws.row_dimensions[row_t].height = 22

            # ── Largeurs colonnes ─────────────────────────────────────────────
            ws.column_dimensions['A'].width = 42
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 16
            for i in range(len(periodes)):
                ws.column_dimensions[get_column_letter(4+i)].width = 14

            # ── Figer les volets ─────────────────────────────────────────────
            ws.freeze_panes = "D3"

            wb.save(path)
            QMessageBox.information(self, "Export réussi",
                f"Tableau de charge exporté :\n{path}")

        except Exception as e:
            logger.error(f"Erreur export ETP: {e}", exc_info=True)
            QMessageBox.critical(self, "Erreur export", str(e))
