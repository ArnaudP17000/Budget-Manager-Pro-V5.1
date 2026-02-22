"""
export_service.py — Export Excel du budget DSI
Amélioration #8
"""
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportService:

    def export_budget_excel(self, exercice, entite_id=None, output_path=None):
        """Export Excel 4 onglets : Synthèse, Lignes, Contrats, BC."""
        # Vérifier openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return False, "openpyxl non installé — lancez : pip install openpyxl"

        # Charger les services
        try:
            from app.services.budget_v5_service import budget_v5_service as bsvc
            from app.services.contrat_service import contrat_service as csvc
            from app.services.bon_commande_service import bon_commande_service as bcsvc
        except Exception as e:
            return False, f"Services indisponibles : {e}"

        try:
            wb  = openpyxl.Workbook()
            now = datetime.now().strftime("%d/%m/%Y %H:%M")
            BLUE = "2980b9"; GREEN = "27ae60"; PURPLE = "8e44ad"
            ORANGE = "e67e22"; RED = "e74c3c"

            def hdr(ws, row, cols, bg="1a252f"):
                fill = PatternFill("solid", fgColor=bg)
                font = Font(bold=True, color="FFFFFF", size=11)
                for c in range(1, cols + 1):
                    ws.cell(row, c).fill = fill
                    ws.cell(row, c).font = font
                    ws.cell(row, c).alignment = Alignment(horizontal="center")

            def colw(ws, widths):
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            def money(v):
                return round(float(v or 0), 2)

            def alt(ws, row, n, even):
                if even:
                    fill = PatternFill("solid", fgColor="f8f9fa")
                    for c in range(1, n + 1):
                        ws.cell(row, c).fill = fill

            # ── Onglet 1 : Synthèse ───────────────────────────────────────
            ws1 = wb.active
            ws1.title = f"Synthèse {exercice}"
            ws1.merge_cells("A1:G1")
            ws1["A1"] = f"SYNTHÈSE BUDGET DSI — Exercice {exercice}"
            ws1["A1"].font = Font(bold=True, size=14, color="FFFFFF")
            ws1["A1"].fill = PatternFill("solid", fgColor="1a252f")
            ws1["A1"].alignment = Alignment(horizontal="center")
            ws1.row_dimensions[1].height = 30
            ws1["A2"] = f"Généré le {now}"
            ws1["A2"].font = Font(italic=True, color="7f8c8d", size=9)

            hdrs1 = ["Entité", "Nature", "Prévisionnel", "Voté", "Engagé", "Solde", "Statut"]
            for i, h in enumerate(hdrs1, 1):
                ws1.cell(4, i, h)
            hdr(ws1, 4, len(hdrs1), BLUE)

            row = 5
            for s in (bsvc.get_synthese_budgets(exercice) or []):
                ws1.cell(row, 1, s.get("entite_nom", ""))
                ws1.cell(row, 2, s.get("nature", ""))
                for c, k in [(3,"montant_previsionnel"),(4,"montant_vote"),
                              (5,"montant_engage"),(6,"montant_solde")]:
                    ws1.cell(row, c, money(s.get(k)))
                    ws1.cell(row, c).number_format = "#,##0.00 €"
                    ws1.cell(row, c).alignment = Alignment(horizontal="right")
                ws1.cell(row, 7, s.get("statut_budget", ""))
                vote = money(s.get("montant_vote"))
                if vote > 0:
                    pct = money(s.get("montant_engage")) / vote * 100
                    col = RED if pct >= 90 else (ORANGE if pct >= 75 else GREEN)
                    ws1.cell(row, 6).font = Font(bold=True, color=col)
                alt(ws1, row, len(hdrs1), row % 2 == 0)
                row += 1
            colw(ws1, [22, 18, 15, 15, 15, 15, 14])

            # ── Onglet 2 : Lignes budgétaires groupées par entité ────────
            ws2 = wb.create_sheet(f"Lignes {exercice}")
            hdrs2 = ["Libellé", "Nature", "Application", "Fournisseur",
                     "Référence D.S.I", "Voté", "Engagé", "Solde", "Taux %", "Alerte"]
            N2 = len(hdrs2)

            # Grouper les lignes par entité puis par nature de budget
            from collections import defaultdict
            lignes_all = bsvc.get_lignes(exercice=exercice) or []
            groupes = defaultdict(list)
            for lb in lignes_all:
                cle = (lb.get('entite_nom','?'), lb.get('budget_nature', lb.get('nature','')), lb.get('exercice', exercice))
                groupes[cle].append(lb)

            row = 1
            for (entite_nom, budget_nature, ex), lignes in sorted(groupes.items()):
                # Titre groupe : Entité — Nature — Exercice
                titre = f"{entite_nom}  —  {budget_nature}  —  {ex}"
                ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N2)
                cell_titre = ws2.cell(row, 1, titre)
                cell_titre.font = Font(bold=True, size=12, color="FFFFFF")
                cell_titre.fill = PatternFill("solid", fgColor=PURPLE)
                cell_titre.alignment = Alignment(horizontal="left", indent=1)
                ws2.row_dimensions[row].height = 22
                row += 1

                # Sous-total entité
                total_vote   = sum(float(lb.get('montant_vote') or 0) for lb in lignes)
                total_engage = sum(float(lb.get('montant_engage') or 0) for lb in lignes)
                total_solde  = sum(float(lb.get('montant_solde') or 0) for lb in lignes)

                # En-têtes colonnes
                for i, h in enumerate(hdrs2, 1):
                    ws2.cell(row, i, h)
                hdr(ws2, row, N2, "2c3e50")
                row += 1

                # Lignes du groupe
                for lb in lignes:
                    ws2.cell(row, 1, lb.get('libelle',''))
                    ws2.cell(row, 2, lb.get('nature',''))
                    ws2.cell(row, 3, lb.get('application_nom') or '')
                    ws2.cell(row, 4, lb.get('fournisseur_nom') or '')
                    ws2.cell(row, 5, lb.get('note') or '')
                    for c, k in [(6,'montant_vote'),(7,'montant_engage'),(8,'montant_solde')]:
                        ws2.cell(row, c, money(lb.get(k)))
                        ws2.cell(row, c).number_format = "#,##0.00 €"
                        ws2.cell(row, c).alignment = Alignment(horizontal="right")
                    taux = float(lb.get('taux_engagement_pct') or 0)
                    ws2.cell(row, 9, taux / 100)
                    ws2.cell(row, 9).number_format = "0.0%"
                    al = lb.get('alerte_seuil', 0)
                    ws2.cell(row, 10, "⚠️ SEUIL" if al else "OK")
                    if al:
                        ws2.cell(row, 10).font = Font(bold=True, color=RED)
                    alt(ws2, row, N2, row % 2 == 0)
                    row += 1

                # Ligne de sous-total
                for c in range(1, N2 + 1):
                    ws2.cell(row, c).fill = PatternFill("solid", fgColor="dfe6e9")
                ws2.cell(row, 1, f"TOTAL  {entite_nom} — {budget_nature}")
                ws2.cell(row, 1).font = Font(bold=True)
                ws2.cell(row, 1).alignment = Alignment(horizontal="right")
                for c, v in [(6, total_vote), (7, total_engage), (8, total_solde)]:
                    ws2.cell(row, c, v)
                    ws2.cell(row, c).number_format = "#,##0.00 €"
                    ws2.cell(row, c).font = Font(bold=True)
                    ws2.cell(row, c).alignment = Alignment(horizontal="right")
                row += 2  # Ligne vide entre groupes

            colw(ws2, [32, 16, 20, 20, 20, 13, 13, 13, 10, 10])

            # ── Onglet 3 : Contrats ───────────────────────────────────────
            ws3 = wb.create_sheet("Contrats actifs")
            hdrs3 = ["Entité","N° Contrat","Type","Objet","Fournisseur","Application",
                     "Date fin","Jours","Montant HT","Montant max","Engagé","Alerte"]
            for i, h in enumerate(hdrs3, 1):
                ws3.cell(1, i, h)
            hdr(ws3, 1, len(hdrs3), "2c3e50")
            row = 2
            for ct in (csvc.get_all() or []):
                ws3.cell(row, 1,  ct.get("entite_code") or "")
                ws3.cell(row, 2,  ct.get("numero_contrat") or "")
                ws3.cell(row, 3,  ct.get("type_contrat") or "")
                ws3.cell(row, 4,  ct.get("objet") or "")
                ws3.cell(row, 5,  ct.get("fournisseur_nom") or "")
                ws3.cell(row, 6,  ct.get("application_nom") or "")
                ws3.cell(row, 7,  str(ct.get("date_fin") or "")[:10])
                ws3.cell(row, 8,  ct.get("jours_restants") or "")
                for c, k in [(9,"montant_ht"),(10,"montant_max_ht"),(11,"montant_engage_cumul")]:
                    ws3.cell(row, c, money(ct.get(k)))
                    ws3.cell(row, c).number_format = "#,##0.00 €"
                    ws3.cell(row, c).alignment = Alignment(horizontal="right")
                niv = ct.get("niveau_alerte","OK")
                ws3.cell(row, 12, niv)
                nc = {"EXPIRE":RED,"CRITIQUE":RED,"ATTENTION":ORANGE,"INFO":BLUE}.get(niv)
                if nc:
                    ws3.cell(row, 12).font = Font(bold=True, color=nc)
                alt(ws3, row, len(hdrs3), row % 2 == 0)
                row += 1
            colw(ws3, [10, 16, 18, 30, 18, 16, 11, 10, 13, 13, 13, 10])

            # ── Onglet 4 : BC ─────────────────────────────────────────────
            ws4 = wb.create_sheet(f"BC {exercice}")
            hdrs4 = ["Entité","N° BC","Date","Fournisseur","Objet","Contrat",
                     "Ligne budgétaire","Application","HT","TTC","Statut"]
            for i, h in enumerate(hdrs4, 1):
                ws4.cell(1, i, h)
            hdr(ws4, 1, len(hdrs4), GREEN)
            row = 2
            filters = {"exercice": exercice} if exercice else None
            for bc in (bcsvc.get_all_bons_commande(filters) or []):
                ws4.cell(row, 1,  bc.get("entite_code") or "")
                ws4.cell(row, 2,  bc.get("numero_bc") or "")
                ws4.cell(row, 3,  str(bc.get("date_creation") or "")[:10])
                ws4.cell(row, 4,  bc.get("fournisseur_nom") or "")
                ws4.cell(row, 5,  bc.get("objet") or "")
                ws4.cell(row, 6,  bc.get("numero_contrat") or "")
                ws4.cell(row, 7,  bc.get("ligne_libelle") or "")
                ws4.cell(row, 8,  bc.get("application_nom") or "")
                for c, k in [(9,"montant_ht"),(10,"montant_ttc")]:
                    ws4.cell(row, c, money(bc.get(k)))
                    ws4.cell(row, c).number_format = "#,##0.00 €"
                    ws4.cell(row, c).alignment = Alignment(horizontal="right")
                ws4.cell(row, 11, bc.get("statut") or "")
                alt(ws4, row, len(hdrs4), row % 2 == 0)
                row += 1
            colw(ws4, [10, 16, 11, 18, 28, 16, 22, 16, 12, 12, 11])

            # ── Onglet 5 : Budget prévisionnel N+1 ───────────────────────
            exercice_n1 = exercice + 1
            ws5 = wb.create_sheet(f"Prévisionnel {exercice_n1}")

            # Titre principal
            ws5.merge_cells("A1:H1")
            ws5["A1"] = f"BUDGET PRÉVISIONNEL {exercice_n1} — DSI"
            ws5["A1"].font = Font(bold=True, size=14, color="FFFFFF")
            ws5["A1"].fill = PatternFill("solid", fgColor=GREEN)
            ws5["A1"].alignment = Alignment(horizontal="center")
            ws5.row_dimensions[1].height = 30

            from collections import defaultdict as _dd

            # Récupérer TOUTES les lignes N+1 et N (sans filtre entité)
            lignes_n1_reelles = bsvc.get_lignes(exercice=exercice_n1) or []
            lignes_n_toutes   = bsvc.get_lignes(exercice=exercice)    or []

            # Récupérer TOUS les budgets annuels N et N+1
            budgets_n_tous  = bsvc.get_budgets(exercice=exercice)    or []
            budgets_n1_tous = bsvc.get_budgets(exercice=exercice_n1) or []

            if lignes_n1_reelles:
                source_label = f"Données réelles {exercice_n1}"
                lignes_base  = lignes_n1_reelles
                use_n1_field = False
            elif lignes_n_toutes:
                source_label = f"Basé sur lignes {exercice} — budget {exercice_n1} non créé"
                lignes_base  = lignes_n_toutes
                use_n1_field = True
            else:
                source_label = f"Basé sur budgets annuels {exercice} — aucune ligne saisie"
                lignes_base  = []
                use_n1_field = False

            ws5["A2"] = f"Généré le {now}  |  Source : {source_label}"
            ws5["A2"].font = Font(italic=True, color="7f8c8d", size=9)

            hdrs5 = ["Entité", "Nature", "Libellé", "Application", "Fournisseur",
                     "Montant prévu N+1", "Référence D.S.I", "Note"]
            N5 = len(hdrs5)
            row = 3
            grand_total = 0

            if lignes_base:
                # Grouper par (entite_nom, nature_budget) — toutes entités confondues
                grp = _dd(list)
                for lb in lignes_base:
                    cle = (lb.get('entite_nom', '?'),
                           lb.get('budget_nature', lb.get('nature', '')))
                    grp[cle].append(lb)

                for (entite_nom, budget_nature), lignes in sorted(grp.items()):
                    # Bandeau groupe
                    ws5.merge_cells(start_row=row, start_column=1,
                                    end_row=row, end_column=N5)
                    cg = ws5.cell(row, 1,
                                  f"{entite_nom}  —  {budget_nature}  —  {exercice_n1}")
                    cg.font = Font(bold=True, size=11, color="FFFFFF")
                    cg.fill = PatternFill("solid", fgColor="27ae60")
                    cg.alignment = Alignment(horizontal="left", indent=1)
                    ws5.row_dimensions[row].height = 20
                    row += 1

                    for i, h in enumerate(hdrs5, 1):
                        ws5.cell(row, i, h)
                    hdr(ws5, row, N5, "2c3e50")
                    row += 1

                    total_grp = 0
                    for lb in lignes:
                        if use_n1_field and float(lb.get('montant_prevu_n1') or 0) > 0:
                            montant = float(lb.get('montant_prevu_n1') or 0)
                        else:
                            montant = float(lb.get('montant_prevu')
                                            or lb.get('montant_vote') or 0)
                        note_src = f"Copié depuis {exercice}" if use_n1_field else ""
                        ws5.cell(row, 1, lb.get('entite_nom', ''))
                        ws5.cell(row, 2, lb.get('nature', ''))
                        ws5.cell(row, 3, lb.get('libelle', ''))
                        ws5.cell(row, 4, lb.get('application_nom') or '')
                        ws5.cell(row, 5, lb.get('fournisseur_nom') or '')
                        ws5.cell(row, 6, montant)
                        ws5.cell(row, 6).number_format = "#,##0.00 €"
                        ws5.cell(row, 6).alignment = Alignment(horizontal="right")
                        ws5.cell(row, 6).font = Font(bold=True, color="27ae60")
                        ws5.cell(row, 7, lb.get('note') or '')
                        ws5.cell(row, 8, note_src)
                        alt(ws5, row, N5, row % 2 == 0)
                        total_grp += montant
                        row += 1

                    for c in range(1, N5 + 1):
                        ws5.cell(row, c).fill = PatternFill("solid", fgColor="d5f5e3")
                    ws5.cell(row, 1, f"SOUS-TOTAL  {entite_nom} — {budget_nature}")
                    ws5.cell(row, 1).font = Font(bold=True)
                    ws5.cell(row, 6, total_grp)
                    ws5.cell(row, 6).number_format = "#,##0.00 €"
                    ws5.cell(row, 6).font = Font(bold=True, color="27ae60")
                    ws5.cell(row, 6).alignment = Alignment(horizontal="right")
                    grand_total += total_grp
                    row += 2

            else:
                # Fallback : pas de lignes → 1 ligne par budget annuel (CDA+Ville × FONCT+INVEST)
                budgets_base = budgets_n1_tous if budgets_n1_tous else budgets_n_tous
                for i, h in enumerate(hdrs5, 1):
                    ws5.cell(row, i, h)
                hdr(ws5, row, N5, "2c3e50")
                row += 1
                for b in sorted(budgets_base,
                                 key=lambda x: (x.get('entite_nom',''), x.get('nature',''))):
                    montant = float(b.get('montant_previsionnel')
                                    or b.get('montant_vote') or 0)
                    ws5.cell(row, 1, b.get('entite_nom', b.get('entite_code', '')))
                    ws5.cell(row, 2, b.get('nature', ''))
                    ws5.cell(row, 3, f"Budget global {b.get('nature','')} {exercice_n1}")
                    ws5.cell(row, 6, montant)
                    ws5.cell(row, 6).number_format = "#,##0.00 €"
                    ws5.cell(row, 6).alignment = Alignment(horizontal="right")
                    ws5.cell(row, 6).font = Font(bold=True, color="27ae60")
                    ws5.cell(row, 8, "Depuis budget annuel")
                    alt(ws5, row, N5, row % 2 == 0)
                    grand_total += montant
                    row += 1

            # Ligne grand total
            ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws5.cell(row, 1, f"TOTAL GÉNÉRAL BUDGET PRÉVISIONNEL {exercice_n1}")
            ws5.cell(row, 1).font = Font(bold=True, size=12, color="FFFFFF")
            ws5.cell(row, 1).fill = PatternFill("solid", fgColor="1a252f")
            ws5.cell(row, 1).alignment = Alignment(horizontal="right")
            ws5.cell(row, 6, grand_total)
            ws5.cell(row, 6).number_format = "#,##0.00 €"
            ws5.cell(row, 6).font = Font(bold=True, size=12, color="FFFFFF")
            ws5.cell(row, 6).fill = PatternFill("solid", fgColor="1a252f")
            ws5.cell(row, 6).alignment = Alignment(horizontal="right")
            ws5.row_dimensions[row].height = 25
            colw(ws5, [22, 16, 32, 20, 20, 16, 22, 14])

            # Sauvegarder
            if not output_path:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                os.makedirs(desktop, exist_ok=True)
                output_path = os.path.join(desktop, f"Budget_DSI_{exercice}_{stamp}.xlsx")

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            logger.info(f"Export Excel : {output_path}")
            return True, output_path

        except Exception as e:
            logger.error(f"Erreur export Excel : {e}", exc_info=True)
            return False, str(e)


export_service = ExportService()
