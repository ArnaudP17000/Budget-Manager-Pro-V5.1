"""
tpe_service.py — CRUD + export + import pour le module TPE.
"""
import json
import logging
import os

logger = logging.getLogger(__name__)


class TpeService:
    def __init__(self):
        from app.services.database_service import DatabaseService
        self.db = DatabaseService()

    # ── Liste ──────────────────────────────────────────────────────────────
    def get_all(self, search=None):
        where, params = [], []
        if search:
            like = '%' + search + '%'
            where.append(
                "(t.service ILIKE %s OR t.regisseur_nom ILIKE %s "
                "OR t.regisseur_prenom ILIKE %s)"
            )
            params += [like, like, like]

        clause = ('WHERE ' + ' AND '.join(where)) if where else ''
        rows = self.db.fetch_all(f"""
            SELECT
                t.*,
                COALESCE(
                    json_agg(
                        json_build_object(
                            'id',              tc.id,
                            'numero',          tc.numero,
                            'numero_serie_tpe',tc.numero_serie_tpe,
                            'modele_tpe',      tc.modele_tpe
                        ) ORDER BY tc.id
                    ) FILTER (WHERE tc.id IS NOT NULL),
                    '[]'
                ) AS cartes
            FROM tpe t
            LEFT JOIN tpe_cartes tc ON tc.tpe_id = t.id
            {clause}
            GROUP BY t.id
            ORDER BY t.service
        """, params)
        return rows or []

    # ── Fiche unique ───────────────────────────────────────────────────────
    def get_by_id(self, tpe_id):
        row = self.db.fetch_one("""
            SELECT
                t.*,
                COALESCE(
                    json_agg(
                        json_build_object(
                            'id',              tc.id,
                            'numero',          tc.numero,
                            'numero_serie_tpe',tc.numero_serie_tpe,
                            'modele_tpe',      tc.modele_tpe
                        ) ORDER BY tc.id
                    ) FILTER (WHERE tc.id IS NOT NULL),
                    '[]'
                ) AS cartes
            FROM tpe t
            LEFT JOIN tpe_cartes tc ON tc.tpe_id = t.id
            WHERE t.id = %s
            GROUP BY t.id
        """, [tpe_id])
        if not row:
            return None
        # cartes peut être une str JSON si le driver ne la désérialise pas
        if isinstance(row.get('cartes'), str):
            try:
                row['cartes'] = json.loads(row['cartes'])
            except Exception:
                row['cartes'] = []
        return {'tpe': row}

    # ── Créer ──────────────────────────────────────────────────────────────
    def create(self, data, created_by_id=None):
        row = self.db.execute_returning("""
            INSERT INTO tpe (
                service, regisseur_prenom, regisseur_nom, regisseur_telephone,
                regisseurs_suppleants, shop_id, backoffice_actif, backoffice_email,
                modele_tpe, type_ethernet, type_4_5g,
                reseau_ip, reseau_masque, reseau_passerelle,
                nombre_tpe, created_by_id, date_maj
            ) VALUES (
                %s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,
                %s,%s,NOW()
            ) RETURNING id
        """, [
            data.get('service'),
            data.get('regisseur_prenom'),
            data.get('regisseur_nom'),
            data.get('regisseur_telephone'),
            data.get('regisseurs_suppleants'),
            data.get('shop_id', 0),
            bool(data.get('backoffice_actif')),
            data.get('backoffice_email'),
            data.get('modele_tpe'),
            bool(data.get('type_ethernet')),
            bool(data.get('type_4_5g')),
            data.get('reseau_ip'),
            data.get('reseau_masque'),
            data.get('reseau_passerelle'),
            data.get('nombre_tpe', 1),
            created_by_id,
        ])
        tpe_id = row[0]
        self._save_cartes(tpe_id, data.get('cartes', []))
        return tpe_id

    # ── Mettre à jour ──────────────────────────────────────────────────────
    def update(self, tpe_id, data):
        self.db.execute("""
            UPDATE tpe SET
                service               = %s,
                regisseur_prenom      = %s,
                regisseur_nom         = %s,
                regisseur_telephone   = %s,
                regisseurs_suppleants = %s,
                shop_id               = %s,
                backoffice_actif      = %s,
                backoffice_email      = %s,
                modele_tpe            = %s,
                type_ethernet         = %s,
                type_4_5g             = %s,
                reseau_ip             = %s,
                reseau_masque         = %s,
                reseau_passerelle     = %s,
                nombre_tpe            = %s,
                date_maj              = NOW()
            WHERE id = %s
        """, [
            data.get('service'),
            data.get('regisseur_prenom'),
            data.get('regisseur_nom'),
            data.get('regisseur_telephone'),
            data.get('regisseurs_suppleants'),
            data.get('shop_id', 0),
            bool(data.get('backoffice_actif')),
            data.get('backoffice_email'),
            data.get('modele_tpe'),
            bool(data.get('type_ethernet')),
            bool(data.get('type_4_5g')),
            data.get('reseau_ip'),
            data.get('reseau_masque'),
            data.get('reseau_passerelle'),
            data.get('nombre_tpe', 1),
            tpe_id,
        ])
        self._save_cartes(tpe_id, data.get('cartes', []))

    # ── Supprimer ──────────────────────────────────────────────────────────
    def delete(self, tpe_id):
        self.db.execute("DELETE FROM tpe WHERE id = %s", [tpe_id])

    # ── Statistiques ───────────────────────────────────────────────────────
    def stats(self):
        row = self.db.fetch_one("""
            SELECT
                COUNT(*)                                    AS total_fiches,
                COALESCE(SUM(nombre_tpe), 0)               AS total_appareils,
                COUNT(*) FILTER (WHERE type_ethernet)      AS nb_ethernet,
                COUNT(*) FILTER (WHERE type_4_5g)          AS nb_4_5g,
                COUNT(*) FILTER (WHERE backoffice_actif)   AS nb_backoffice
            FROM tpe
        """)
        return {
            'total_fiches':    int(row['total_fiches']    or 0),
            'total_appareils': int(row['total_appareils'] or 0),
            'nb_ethernet':     int(row['nb_ethernet']     or 0),
            'nb_4_5g':         int(row['nb_4_5g']         or 0),
            'nb_backoffice':   int(row['nb_backoffice']   or 0),
        }

    # ── Export Excel ───────────────────────────────────────────────────────
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl non installé (pip install openpyxl)")

        rows = self.get_all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TPE"

        headers = [
            'Service', 'Régisseur Prénom', 'Régisseur Nom', 'Téléphone',
            'Régisseurs suppléants', 'Shop ID', 'Modèle TPE',
            'Ethernet', '4/5G', 'IP', 'Masque', 'Passerelle',
            'Backoffice actif', 'Email backoffice', 'Nb TPE',
            'Cartes (numéros)',
        ]
        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for r, t in enumerate(rows, 2):
            cartes_nums = ', '.join(
                c['numero'] for c in (t.get('cartes') or []) if c.get('numero')
            )
            ws.append([
                t.get('service'),
                t.get('regisseur_prenom'),
                t.get('regisseur_nom'),
                t.get('regisseur_telephone'),
                t.get('regisseurs_suppleants'),
                t.get('shop_id'),
                t.get('modele_tpe'),
                'Oui' if t.get('type_ethernet') else 'Non',
                'Oui' if t.get('type_4_5g')    else 'Non',
                t.get('reseau_ip'),
                t.get('reseau_masque'),
                t.get('reseau_passerelle'),
                'Oui' if t.get('backoffice_actif') else 'Non',
                t.get('backoffice_email'),
                t.get('nombre_tpe', 1),
                cartes_nums,
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Import JSON (idempotent) ───────────────────────────────────────────
    def import_from_json(self, path):
        count = self.db.fetch_one("SELECT COUNT(*) AS n FROM tpe")
        if count and int(count['n']) > 0:
            return 0  # déjà peuplé

        with open(path, encoding='utf-8') as f:
            raw = json.load(f)

        # Support format imbriqué (TpeComplet-v2) ou format plat
        records = raw.get('tpes', raw) if isinstance(raw, dict) else raw

        imported = 0
        for rec in records:
            try:
                flat = self._normalize_import(rec)
                self.create(flat)
                imported += 1
            except Exception as e:
                logger.warning("TPE import skipped: %s", e)
        return imported

    @staticmethod
    def _normalize_import(rec):
        """Convertit le format imbriqué TpeComplet-v2 vers le format plat."""
        # Déjà plat (champs directs)
        if 'regisseur_nom' in rec or 'regisseur_prenom' in rec:
            return rec

        reg  = rec.get('regisseur') or {}
        bo   = rec.get('acces_backoffice') or {}
        ttyp = rec.get('type_tpe') or {}
        net  = ttyp.get('config_reseau') or {}

        cartes_src = rec.get('cartes_commercant') or rec.get('cartes') or []
        cartes = [
            {
                'numero':          str(c.get('numero', '')),
                'numero_serie_tpe': c.get('numero_serie_tpe'),
                'modele_tpe':       c.get('modele_tpe'),
            }
            for c in cartes_src if c.get('numero')
        ]

        return {
            'service':               rec.get('service'),
            'regisseur_prenom':      reg.get('prenom'),
            'regisseur_nom':         reg.get('nom'),
            'regisseur_telephone':   reg.get('telephone'),
            'regisseurs_suppleants': rec.get('regisseurs_suppleants'),
            'shop_id':               rec.get('shop_id', 0),
            'backoffice_actif':      bool(bo.get('actif')),
            'backoffice_email':      bo.get('email'),
            'modele_tpe':            rec.get('modele_tpe'),
            'type_ethernet':         bool(ttyp.get('ethernet')),
            'type_4_5g':             bool(ttyp.get('quatre_cinq_g')),
            'reseau_ip':             net.get('adresse_ip'),
            'reseau_masque':         net.get('masque'),
            'reseau_passerelle':     net.get('passerelle'),
            'nombre_tpe':            rec.get('nombre_tpe', 1),
            'cartes':                cartes,
        }

    # ── Cartes (helper interne) ────────────────────────────────────────────
    def _save_cartes(self, tpe_id, cartes):
        self.db.execute("DELETE FROM tpe_cartes WHERE tpe_id = %s", [tpe_id])
        for c in (cartes or []):
            numero = (c.get('numero') or '').strip()
            if not numero:
                continue
            self.db.execute(
                "INSERT INTO tpe_cartes (tpe_id, numero, numero_serie_tpe, modele_tpe) "
                "VALUES (%s, %s, %s, %s)",
                [tpe_id, numero, c.get('numero_serie_tpe'), c.get('modele_tpe')]
            )
