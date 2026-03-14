import functools
import hashlib
import hmac
import json
import secrets
import time
from collections import defaultdict
from urllib.parse import urlencode
from flask import Blueprint, jsonify, request, g, redirect as flask_redirect

from app.services.projet_service import ProjetService
from app.services.budget_v5_service import BudgetV5Service
from app.services.bon_commande_service import BonCommandeService
from app.services.contrat_service import ContratService
from app.services.tache_service import TacheService
from app.services.referentiel_service import ReferentielService
from app.services.contact_service import ContactService
from app.services.service_org_service import ServiceOrgService
from app.services.auth_service import AuthService

routes = Blueprint('routes', __name__)

# ── Helpers réponses standardisées ──────────────────────────────────────────

def _err(msg, code=400):
    """Réponse d'erreur uniforme : {"success": false, "error": "..."}"""
    return jsonify({"success": False, "error": str(msg)}), code

def _ok(**data):
    """Réponse de succès uniforme : {"success": true, ...}"""
    return jsonify({"success": True, **data})

# ── Validation des inputs ────────────────────────────────────────────────────

def _validate(data, rules):
    """
    Valide un dict de données.
    rules = {
        'field': {
            'label':    str   (affiché dans l'erreur, défaut = nom du champ),
            'required': bool,
            'max':      int   (longueur max),
            'type':     'number'|'email'|'date',
            'enum':     list[str],
        }
    }
    Retourne une liste de messages d'erreur (vide = valide).
    """
    errors = []
    for field, rule in rules.items():
        val  = data.get(field)
        lbl  = rule.get('label', field)
        sval = str(val).strip() if val is not None else ''

        if rule.get('required') and not sval:
            errors.append(f"'{lbl}' est obligatoire")
            continue

        if not sval:
            continue  # champ vide non requis — ok

        if rule.get('max') and len(sval) > rule['max']:
            errors.append(f"'{lbl}' dépasse {rule['max']} caractères")

        if rule.get('type') == 'number':
            try:
                float(val)
            except (TypeError, ValueError):
                errors.append(f"'{lbl}' doit être un nombre valide")

        if rule.get('type') == 'email' and '@' not in sval:
            errors.append(f"'{lbl}' n'est pas une adresse email valide")

        if rule.get('enum') and sval not in rule['enum']:
            errors.append(f"'{lbl}' : valeur invalide (attendu : {', '.join(rule['enum'])})")

    return errors

projet_service      = ProjetService()
budget_service      = BudgetV5Service()
bc_service          = BonCommandeService()
contrat_service     = ContratService()
referentiel_service = ReferentielService()
tache_service       = TacheService()
contact_service     = ContactService()
service_org_service = ServiceOrgService()
auth_service        = AuthService()


# ─────────────────────────────────────────────
# DECORATOR AUTH
# ─────────────────────────────────────────────

def require_auth(*roles):
    """
    @require_auth()                      → tout utilisateur authentifié
    @require_auth('admin')               → admin seulement
    @require_auth('admin','gestionnaire')→ admin ou gestionnaire
    """
    def decorator(f):
        @functools.wraps(f)
        def wrapper(*args, **kwargs):
            auth_header = request.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                return jsonify({"error": "Token manquant"}), 401
            payload = auth_service.verify_token(auth_header[7:])
            if not payload:
                return jsonify({"error": "Token invalide ou expiré"}), 401
            if roles and payload.get('role') not in roles:
                return jsonify({"error": "Accès interdit"}), 403
            # Vérifier que le compte est toujours actif en base
            try:
                user_row = auth_service.db.fetch_one(
                    "SELECT actif FROM utilisateurs WHERE id=%s", [payload.get('sub')]
                )
                if not user_row or not user_row.get('actif'):
                    return jsonify({"error": "Compte désactivé"}), 401
            except Exception:
                pass  # DB temporairement indisponible : on accepte le token JWT
            g.user = payload
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────

# Rate limiting — max 10 tentatives par fenêtre de 5 min par IP
_login_attempts: dict = defaultdict(list)
_LOGIN_WINDOW   = 300   # secondes
_LOGIN_MAX      = 10    # tentatives max

def _check_login_rate(ip: str) -> bool:
    now = time.time()
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOGIN_WINDOW]
    if len(_login_attempts[ip]) >= _LOGIN_MAX:
        return False
    _login_attempts[ip].append(now)
    return True

@routes.route('/auth/login', methods=['POST'])
def login():
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    if not _check_login_rate(ip):
        return jsonify({"error": "Trop de tentatives, réessayez dans quelques minutes"}), 429
    data = request.json or {}
    result = auth_service.login(data.get('login', ''), data.get('password', ''))
    if not result:
        return jsonify({"error": "Identifiants invalides ou compte désactivé"}), 401
    return jsonify(result)

@routes.route('/auth/me', methods=['GET'])
@require_auth()
def me():
    return jsonify(g.user)

@routes.route('/auth/refresh', methods=['POST'])
@require_auth()
def refresh_token():
    """Émet un nouveau token JWT si le token actuel est encore valide."""
    from app.services.auth_service import AuthService, SECRET_KEY, EXPIRY_HOURS
    import jwt as _jwt
    from datetime import datetime, timezone, timedelta
    u = g.user
    payload = {
        'sub':        str(u['sub']),
        'login':      u.get('login'),
        'nom':        u.get('nom'),
        'prenom':     u.get('prenom'),
        'role':       u.get('role'),
        'service_id': u.get('service_id'),
        'iat':        datetime.now(timezone.utc),
        'exp':        datetime.now(timezone.utc) + timedelta(hours=EXPIRY_HOURS),
    }
    new_token = _jwt.encode(payload, SECRET_KEY, algorithm='HS256')
    return jsonify({'token': new_token})


# ─────────────────────────────────────────────
# USERS (admin)
# ─────────────────────────────────────────────

@routes.route('/users', methods=['GET'])
@require_auth('admin')
def get_users():
    return jsonify({"list": auth_service.get_all_users()})

@routes.route('/users', methods=['POST'])
@require_auth('admin')
def create_user():
    data = request.json or {}
    try:
        if not data.get('login') or not data.get('mot_de_passe'):
            return jsonify({"error": "login et mot_de_passe requis"}), 400
        auth_service.create_user(data)
        _audit('CREATE', 'utilisateurs', None, {'login': data.get('login'), 'role': data.get('role')})
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@routes.route('/users/<int:user_id>', methods=['GET'])
@require_auth('admin')
def get_user(user_id):
    user = auth_service.get_user_by_id(user_id)
    if not user:
        return jsonify({"error": "Utilisateur introuvable"}), 404
    return jsonify(user)

@routes.route('/users/<int:user_id>', methods=['PUT'])
@require_auth('admin')
def update_user(user_id):
    data = request.json or {}
    try:
        auth_service.update_user(user_id, data)
        _audit('UPDATE', 'utilisateurs', user_id, {'role': data.get('role')})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@routes.route('/users/<int:user_id>', methods=['DELETE'])
@require_auth('admin')
def delete_user(user_id):
    try:
        if g.user.get('sub') == user_id:
            return jsonify({"error": "Impossible de supprimer son propre compte"}), 400
        auth_service.delete_user(user_id)
        _audit('DELETE', 'utilisateurs', user_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@routes.route('/users/<int:user_id>/toggle', methods=['POST'])
@require_auth('admin')
def toggle_user(user_id):
    data = request.json or {}
    try:
        auth_service.set_active(user_id, data.get('actif', True))
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─────────────────────────────────────────────
# JOURNAL D'AUDIT
# ─────────────────────────────────────────────

@routes.route('/audit_log', methods=['GET'])
@require_auth('admin')
def get_audit_log():
    table  = request.args.get('table', '')
    action = request.args.get('action', '')
    limit  = min(int(request.args.get('limit', 200)), 500)
    query  = (
        "SELECT al.id, al.user_id, al.user_login, al.action, al.table_name, "
        "al.record_id, al.details, al.date_creation, "
        "u.nom, u.prenom FROM audit_log al "
        "LEFT JOIN utilisateurs u ON u.id = al.user_id WHERE 1=1"
    )
    params = []
    if table:  query += " AND al.table_name = %s"; params.append(table)
    if action: query += " AND al.action = %s";     params.append(action)
    query += " ORDER BY al.date_creation DESC LIMIT %s"
    params.append(limit)
    rows = bc_service.db.fetch_all(query, params)
    return jsonify({"list": [dict(r) for r in rows] if rows else []})


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@routes.route('/dashboard', methods=['GET'])
@require_auth()
def dashboard():
    try: projets = projet_service.get_all()
    except Exception: projets = []
    try: budget = budget_service.get_budget()
    except Exception: budget = []
    try: bons_commande = bc_service.get_all_bons_commande()
    except Exception: bons_commande = []
    try: contrats = contrat_service.get_all()
    except Exception: contrats = []
    try: alertes = contrat_service.get_alertes()
    except Exception: alertes = []

    bc_attente = [b for b in bons_commande if b.get('statut') in ('BROUILLON', 'EN_ATTENTE')]

    # Lignes budgétaires en alerte (taux_engagement >= 80%)
    try:
        row_alerte = bc_service.db.fetch_one(
            "SELECT COUNT(*) as cnt FROM lignes_budgetaires "
            "WHERE montant_vote > 0 AND montant_engage::numeric * 100 / montant_vote::numeric >= 80"
        )
        kpi_alertes_lignes = int(row_alerte['cnt']) if row_alerte else 0
    except Exception:
        kpi_alertes_lignes = 0

    # Répartition par nature (pour doughnut chart)
    nature_map = {}
    for b in budget:
        n = b.get('nature') or 'AUTRE'
        if n not in nature_map:
            nature_map[n] = {'vote': 0, 'engage': 0}
        nature_map[n]['vote']   += float(b.get('montant_vote', 0) or 0)
        nature_map[n]['engage'] += float(b.get('montant_engage', 0) or 0)

    # Engagement par entité (pour bar chart)
    entite_map = {}
    for b in budget:
        e = b.get('entite_nom') or b.get('entite_code') or 'Inconnu'
        if e not in entite_map:
            entite_map[e] = {'vote': 0, 'engage': 0}
        entite_map[e]['vote']   += float(b.get('montant_vote', 0) or 0)
        entite_map[e]['engage'] += float(b.get('montant_engage', 0) or 0)

    return jsonify({
        "kpi_projets":        len([p for p in projets if p.get('statut') == 'ACTIF']),
        "kpi_budget":         sum(b.get('montant_vote', 0) or 0 for b in budget),
        "kpi_bons_commande":  len(bons_commande),
        "kpi_montant_bc":     sum(b.get('montant_ttc', 0) or 0 for b in bons_commande),
        "kpi_contrats":       len([c for c in contrats if c.get('statut') == 'ACTIF']),
        "kpi_alertes_contrats": len(alertes),
        "kpi_bc_attente":     len(bc_attente),
        "alertes_contrats":   alertes[:5],
        "kpi_alertes_lignes": kpi_alertes_lignes,
        "repartition_nature": [{"nature": k, "vote": v['vote'], "engage": v['engage']}
                                for k, v in nature_map.items()],
        "engagement_entite":  sorted(
            [{"entite": k, "vote": v['vote'], "engage": v['engage']}
             for k, v in entite_map.items()],
            key=lambda x: x['engage'], reverse=True
        )[:10],
    })


# ─────────────────────────────────────────────
# ENTITÉS
# ─────────────────────────────────────────────

@routes.route('/entites', methods=['GET'])
@require_auth()
def get_entites():
    return jsonify({"list": budget_service.get_entites()})


# ─────────────────────────────────────────────
# APPLICATIONS
# ─────────────────────────────────────────────

@routes.route('/applications', methods=['GET'])
@require_auth()
def get_applications():
    return jsonify({"list": budget_service.get_all_applications()})


# ─────────────────────────────────────────────
# BUDGETS ANNUELS — avec contrôle d'accès explicite
# ─────────────────────────────────────────────

def _user_budget_role(user_id, budget_id):
    """Retourne le rôle de l'utilisateur sur ce budget ('lecteur'/'gestionnaire'), ou None."""
    row = budget_service.db.fetch_one(
        "SELECT role FROM budget_permissions WHERE budget_id=%s AND user_id=%s",
        [budget_id, user_id]
    )
    return row['role'] if row else None


@routes.route('/budget', methods=['GET'])
@require_auth()
def get_budget():
    user_id  = g.user.get('sub')
    role     = g.user.get('role')
    exercice = request.args.get('exercice', type=int)

    if role in ('admin', 'gestionnaire_service'):
        budgets = budget_service.get_budget()
        perm = 'gestionnaire' if role == 'admin' else 'lecteur'
        for b in budgets:
            b['user_perm'] = perm
    else:
        rows = budget_service.db.fetch_all(
            "SELECT ba.*, e.code as entite_code, e.nom as entite_nom, bp.role as user_perm "
            "FROM budgets_annuels ba "
            "LEFT JOIN entites e ON e.id = ba.entite_id "
            "JOIN budget_permissions bp ON bp.budget_id = ba.id AND bp.user_id = %s "
            "ORDER BY ba.exercice DESC, ba.nature",
            [user_id]
        )
        budgets = [dict(b) for b in (rows or [])]

    if exercice:
        budgets = [b for b in budgets if b.get('exercice') == exercice]

    return jsonify({
        "total_vote":   sum(b.get('montant_vote', 0) or 0 for b in budgets),
        "total_engage": sum(b.get('montant_engage', 0) or 0 for b in budgets),
        "details": budgets
    })


@routes.route('/budget/<int:budget_id>/lignes', methods=['GET'])
@require_auth()
def get_lignes_budget(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role not in ('admin', 'gestionnaire_service') and _user_budget_role(user_id, budget_id) is None:
        return jsonify({"error": "Accès interdit à ce budget"}), 403
    return jsonify({"list": budget_service.get_lignes(budget_id)})


@routes.route('/lignes', methods=['GET'])
@require_auth()
def get_all_lignes():
    user_id   = g.user.get('sub')
    role      = g.user.get('role')
    budget_id = request.args.get('budget_id', type=int)
    if budget_id:
        if role not in ('admin', 'gestionnaire_service') and _user_budget_role(user_id, budget_id) is None:
            return jsonify({"list": []})
        return jsonify({"list": budget_service.get_lignes(budget_id)})
    if role in ('admin', 'gestionnaire_service'):
        rows = budget_service.db.fetch_all(
            "SELECT l.*, ba.exercice, ba.nature, "
            "e.code as entite_code, e.nom as entite_nom, "
            "CONCAT(e.code, ' — ', ba.nature, ' ', ba.exercice) as budget_label, "
            "app.nom as application_nom, f.nom as fournisseur_nom, "
            "CASE WHEN l.montant_vote > 0 "
            "THEN ROUND((COALESCE(l.montant_engage,0)/l.montant_vote*100)::numeric, 1) "
            "ELSE 0 END as taux_engagement, "
            "CASE WHEN l.montant_vote > 0 AND COALESCE(l.montant_engage,0) >= l.montant_vote * "
            "COALESCE(l.seuil_alerte_pct,80)/100.0 THEN true ELSE false END as alerte "
            "FROM lignes_budgetaires l "
            "JOIN budgets_annuels ba ON ba.id = l.budget_id "
            "LEFT JOIN entites e ON e.id = ba.entite_id "
            "LEFT JOIN applications app ON app.id = l.application_id "
            "LEFT JOIN fournisseurs f ON f.id = l.fournisseur_id "
            "ORDER BY l.id"
        )
    else:
        rows = budget_service.db.fetch_all(
            "SELECT l.*, ba.exercice, ba.nature, "
            "e.code as entite_code, e.nom as entite_nom, "
            "CONCAT(e.code, ' — ', ba.nature, ' ', ba.exercice) as budget_label, "
            "app.nom as application_nom, f.nom as fournisseur_nom, "
            "CASE WHEN l.montant_vote > 0 "
            "THEN ROUND((COALESCE(l.montant_engage,0)/l.montant_vote*100)::numeric, 1) "
            "ELSE 0 END as taux_engagement, "
            "CASE WHEN l.montant_vote > 0 AND COALESCE(l.montant_engage,0) >= l.montant_vote * "
            "COALESCE(l.seuil_alerte_pct,80)/100.0 THEN true ELSE false END as alerte "
            "FROM lignes_budgetaires l "
            "JOIN budgets_annuels ba ON ba.id = l.budget_id "
            "LEFT JOIN entites e ON e.id = ba.entite_id "
            "LEFT JOIN applications app ON app.id = l.application_id "
            "LEFT JOIN fournisseurs f ON f.id = l.fournisseur_id "
            "JOIN budget_permissions bp ON bp.budget_id = l.budget_id AND bp.user_id = %s "
            "ORDER BY l.id",
            [user_id]
        )
    return jsonify({"list": [dict(r) for r in (rows or [])]})


@routes.route('/ligne', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_ligne():
    data    = request.json
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    bid     = data.get('budget_id')
    if role != 'admin' and _user_budget_role(user_id, bid) != 'gestionnaire':
        return jsonify({"error": "Droits insuffisants sur ce budget"}), 403
    try:
        vote = float(data.get('montant_vote') or 0)
        budget_service.db.execute(
            "INSERT INTO lignes_budgetaires "
            "(budget_id, libelle, application_id, fournisseur_id, "
            "montant_prevu, montant_vote, montant_solde, nature, note, statut) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            [bid, data.get('libelle'),
             data.get('application_id') or None, data.get('fournisseur_id') or None,
             float(data.get('montant_prevu') or 0), vote, vote,
             data.get('nature') or 'FONCTIONNEMENT',
             data.get('note') or None, data.get('statut') or 'ACTIF']
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/ligne/<int:ligne_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_ligne(ligne_id):
    data    = request.json
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    bid     = data.get('budget_id')
    if role != 'admin' and _user_budget_role(user_id, bid) != 'gestionnaire':
        return jsonify({"error": "Droits insuffisants sur ce budget"}), 403
    try:
        vote = float(data.get('montant_vote') or 0)
        budget_service.db.execute(
            "UPDATE lignes_budgetaires SET "
            "budget_id=%s, libelle=%s, application_id=%s, fournisseur_id=%s, "
            "montant_prevu=%s, montant_vote=%s, "
            "montant_solde=GREATEST(%s - COALESCE(montant_engage, 0), 0), "
            "nature=%s, note=%s, statut=%s, date_maj=NOW() WHERE id=%s",
            [bid, data.get('libelle'),
             data.get('application_id') or None, data.get('fournisseur_id') or None,
             float(data.get('montant_prevu') or 0), vote, vote,
             data.get('nature') or 'FONCTIONNEMENT',
             data.get('note') or None, data.get('statut') or 'ACTIF', ligne_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/ligne/<int:ligne_id>/bcs', methods=['GET'])
@require_auth()
def get_ligne_bcs(ligne_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin':
        row = budget_service.db.fetch_one(
            "SELECT l.budget_id FROM lignes_budgetaires l WHERE l.id=%s", [ligne_id]
        )
        if not row or _user_budget_role(user_id, row['budget_id']) is None:
            return jsonify({"bcs": []})
    bcs = budget_service.db.fetch_all(
        "SELECT bc.id, bc.numero_bc, bc.objet, bc.montant_ht, bc.montant_ttc, "
        "bc.statut, bc.date_creation, bc.date_imputation, bc.date_solde, "
        "f.nom as fournisseur_nom, "
        "c.numero_contrat as contrat_numero, c.objet as contrat_objet, "
        "p.nom as projet_nom "
        "FROM bons_commande bc "
        "LEFT JOIN fournisseurs f ON f.id = bc.fournisseur_id "
        "LEFT JOIN contrats c ON c.id = bc.contrat_id "
        "LEFT JOIN projets p ON p.id = bc.projet_id "
        "WHERE bc.ligne_budgetaire_id = %s ORDER BY bc.date_creation DESC",
        [ligne_id]
    )
    return jsonify({"bcs": [dict(b) for b in bcs] if bcs else []})


@routes.route('/budget/<int:budget_id>/detail', methods=['GET'])
@require_auth()
def get_budget_detail(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) is None:
        return jsonify({"error": "Accès interdit à ce budget"}), 403
    lignes = budget_service.get_lignes(budget_id)
    for ligne in lignes:
        bcs = budget_service.db.fetch_all(
            "SELECT bc.id, bc.numero_bc, bc.objet, bc.montant_ht, bc.montant_ttc, "
            "bc.statut, bc.date_creation, f.nom as fournisseur_nom, "
            "c.numero_contrat as contrat_numero, c.objet as contrat_objet "
            "FROM bons_commande bc "
            "LEFT JOIN fournisseurs f ON f.id = bc.fournisseur_id "
            "LEFT JOIN contrats c ON c.id = bc.contrat_id "
            "WHERE bc.ligne_budgetaire_id = %s ORDER BY bc.date_creation DESC",
            [ligne['id']]
        )
        ligne['bons_commande'] = [dict(b) for b in bcs] if bcs else []
    return jsonify({"lignes": lignes})


@routes.route('/budget/<int:budget_id>/voter', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def voter_budget(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) != 'gestionnaire':
        return jsonify({"error": "Droits insuffisants sur ce budget"}), 403
    data = request.json
    try:
        montant_vote = float(data.get('montant_vote') or 0)
        budget_service.voter_budget(budget_id, montant_vote)
        _audit('VOTER', 'budgets_annuels', budget_id, {'montant_vote': montant_vote})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/budget', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_budget():
    data    = request.json
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    try:
        row = budget_service.db.execute_returning(
            "INSERT INTO budgets_annuels (entite_id, exercice, nature, montant_previsionnel, statut) "
            "VALUES (%s, %s, %s, %s, %s) RETURNING id",
            [data.get('entite_id') or None, data.get('exercice'), data.get('nature'),
             data.get('montant_previsionnel') or 0, data.get('statut', 'BROUILLON')]
        )
        new_id = row[0] if row else None
        if new_id and role != 'admin':
            budget_service.db.execute(
                "INSERT INTO budget_permissions (budget_id, user_id, role) "
                "VALUES (%s, %s, 'gestionnaire') ON CONFLICT DO NOTHING",
                [new_id, user_id]
            )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/budget/<int:budget_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_budget(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) != 'gestionnaire':
        return jsonify({"error": "Droits insuffisants sur ce budget"}), 403
    data = request.json
    try:
        budget_service.db.execute(
            "UPDATE budgets_annuels SET entite_id=%s, exercice=%s, nature=%s, "
            "montant_previsionnel=%s, montant_vote=%s, statut=%s, date_maj=NOW() WHERE id=%s",
            [data.get('entite_id') or None, data.get('exercice'), data.get('nature'),
             data.get('montant_previsionnel') or 0, data.get('montant_vote') or 0,
             data.get('statut'), budget_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/budget/<int:budget_id>', methods=['DELETE'])
@require_auth('admin')
def delete_budget(budget_id):
    try:
        budget_service.db.execute("DELETE FROM budgets_annuels WHERE id=%s", [budget_id])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─── Simulation + Duplication budget N → N+1 ──────────────────────────────────

@routes.route('/budget/preview_n1', methods=['GET'])
@require_auth('admin')
def preview_budget_n1():
    """Retourne la simulation N+1 sans rien écrire en base."""
    source_exercice = request.args.get('source_exercice', type=int)
    taux = float(request.args.get('taux', 3.5))
    if not source_exercice:
        return jsonify({'error': 'source_exercice requis'}), 400
    coeff = 1 + taux / 100
    budgets = budget_service.db.fetch_all(
        "SELECT ba.*, e.code as entite_code, e.nom as entite_nom "
        "FROM budgets_annuels ba LEFT JOIN entites e ON e.id = ba.entite_id "
        "WHERE ba.exercice=%s ORDER BY ba.id", [source_exercice]
    ) or []
    if not budgets:
        return jsonify({'error': f'Aucun budget pour {source_exercice}'}), 404
    result = []
    for b in budgets:
        engage = float(b.get('montant_engage') or 0)
        vote   = float(b.get('montant_vote')   or 0)
        base   = engage if engage > 0 else vote
        lignes = budget_service.db.fetch_all(
            "SELECT l.*, app.nom as application_nom, f.nom as fournisseur_nom "
            "FROM lignes_budgetaires l "
            "LEFT JOIN applications app ON app.id = l.application_id "
            "LEFT JOIN fournisseurs f ON f.id = l.fournisseur_id "
            "WHERE l.budget_id=%s AND l.statut != 'ANNULEE' ORDER BY l.id",
            [b['id']]
        ) or []
        lignes_sim = []
        for l in lignes:
            eng_l  = float(l.get('montant_engage') or 0)
            vot_l  = float(l.get('montant_vote')   or 0)
            base_l = eng_l if eng_l > 0 else vot_l
            lignes_sim.append({
                'libelle':         l['libelle'],
                'application_nom': l.get('application_nom'),
                'fournisseur_nom': l.get('fournisseur_nom'),
                'nature':          l.get('nature'),
                'base_n':          round(base_l, 2),
                'prevu_n1':        round(base_l * coeff, 2),
            })
        result.append({
            'budget_id':      b['id'],
            'entite_code':    b.get('entite_code'),
            'entite_nom':     b.get('entite_nom'),
            'nature':         b['nature'],
            'base_n':         round(base, 2),
            'previsionnel_n1': round(base * coeff, 2),
            'lignes':         lignes_sim,
        })
    return jsonify({
        'source_exercice': source_exercice,
        'target_exercice': source_exercice + 1,
        'taux':            taux,
        'budgets':         result,
        'total_n':         round(sum(r['base_n']          for r in result), 2),
        'total_n1':        round(sum(r['previsionnel_n1'] for r in result), 2),
    })


# ─── Duplication budget N → N+1 ───────────────────────────────────────────────

@routes.route('/budget/dupliquer', methods=['POST'])
@require_auth('admin')
def dupliquer_budget():
    data            = request.json
    source_exercice = data.get('source_exercice')
    target_exercice = data.get('target_exercice')
    taux            = float(data.get('taux_revalorisation') or 0)
    coeff           = 1 + taux / 100
    if not source_exercice or not target_exercice:
        return jsonify({"error": "source_exercice et target_exercice requis"}), 400
    try:
        existing = budget_service.db.fetch_one(
            "SELECT COUNT(*) as n FROM budgets_annuels WHERE exercice=%s", [target_exercice]
        )
        if existing and int(existing['n'] or 0) > 0:
            return jsonify({"error": f"Des budgets {target_exercice} existent déjà ({existing['n']}). Supprimez-les avant de dupliquer."}), 400

        source_budgets = budget_service.db.fetch_all(
            "SELECT * FROM budgets_annuels WHERE exercice=%s ORDER BY id", [source_exercice]
        ) or []
        if not source_budgets:
            return jsonify({"error": f"Aucun budget trouvé pour l'exercice {source_exercice}"}), 404

        nb_budgets = 0
        nb_lignes  = 0
        for b in source_budgets:
            engage = float(b.get('montant_engage') or 0)
            vote   = float(b.get('montant_vote') or 0)
            base   = engage if engage > 0 else vote
            previsionnel = round(base * coeff, 2)

            new_budget_row = budget_service.db.execute_returning(
                "INSERT INTO budgets_annuels (entite_id, exercice, nature, montant_previsionnel, statut) "
                "VALUES (%s, %s, %s, %s, 'BROUILLON') RETURNING id",
                [b['entite_id'], target_exercice, b['nature'], previsionnel]
            )
            new_budget_id = new_budget_row[0]
            nb_budgets += 1

            lignes = budget_service.db.fetch_all(
                "SELECT * FROM lignes_budgetaires WHERE budget_id=%s AND statut != 'ANNULEE' ORDER BY id",
                [b['id']]
            ) or []
            for l in lignes:
                engage_l = float(l.get('montant_engage') or 0)
                vote_l   = float(l.get('montant_vote') or 0)
                base_l   = engage_l if engage_l > 0 else vote_l
                prevu    = round(base_l * coeff, 2)
                budget_service.db.execute(
                    "INSERT INTO lignes_budgetaires "
                    "(budget_id, libelle, application_id, fournisseur_id, "
                    "montant_prevu, montant_vote, montant_solde, nature, note, statut) "
                    "VALUES (%s, %s, %s, %s, %s, 0, 0, %s, %s, 'ACTIF')",
                    [new_budget_id, l['libelle'],
                     l.get('application_id'), l.get('fournisseur_id'),
                     prevu, l.get('nature') or 'FONCTIONNEMENT', l.get('note')]
                )
                nb_lignes += 1

        return jsonify({"success": True, "budgets_crees": nb_budgets, "lignes_creees": nb_lignes})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── Gestion des permissions budget ───────────────────────────────────────────

@routes.route('/budget/<int:budget_id>/permissions', methods=['GET'])
@require_auth()
def get_budget_permissions(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) != 'gestionnaire':
        return jsonify({"error": "Accès interdit"}), 403
    rows = budget_service.db.fetch_all(
        "SELECT bp.id, bp.user_id, bp.role, bp.date_creation, "
        "u.nom, u.prenom, u.login, s.nom as service_nom "
        "FROM budget_permissions bp "
        "JOIN utilisateurs u ON u.id = bp.user_id "
        "LEFT JOIN services s ON s.id = u.service_id "
        "WHERE bp.budget_id = %s ORDER BY u.nom, u.prenom",
        [budget_id]
    )
    return jsonify({"list": [dict(r) for r in (rows or [])]})


@routes.route('/budget/<int:budget_id>/permissions', methods=['POST'])
@require_auth()
def add_budget_permission(budget_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) != 'gestionnaire':
        return jsonify({"error": "Accès interdit"}), 403
    data = request.json or {}
    perm_role = data.get('role', 'lecteur')
    if perm_role not in ('lecteur', 'gestionnaire'):
        return jsonify({"error": "Rôle invalide — valeurs acceptées : lecteur, gestionnaire"}), 400
    target_uid = data.get('user_id')
    if not target_uid or not str(target_uid).isdigit():
        return jsonify({"error": "user_id invalide"}), 400
    try:
        budget_service.db.execute(
            "INSERT INTO budget_permissions (budget_id, user_id, role) VALUES (%s, %s, %s) "
            "ON CONFLICT (budget_id, user_id) DO UPDATE SET role = EXCLUDED.role",
            [budget_id, int(target_uid), perm_role]
        )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/budget/<int:budget_id>/permissions/<int:target_uid>', methods=['DELETE'])
@require_auth()
def delete_budget_permission(budget_id, target_uid):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    if role != 'admin' and _user_budget_role(user_id, budget_id) != 'gestionnaire':
        return jsonify({"error": "Accès interdit"}), 403
    try:
        budget_service.db.execute(
            "DELETE FROM budget_permissions WHERE budget_id=%s AND user_id=%s",
            [budget_id, target_uid]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/budget/vue_service', methods=['GET'])
@require_auth('gestionnaire_service')
def get_budget_vue_service():
    """Vue agrégée des accès budgets par membre du service (gestionnaire_service uniquement)."""
    service_id = g.user.get('service_id')
    if not service_id:
        return jsonify({'users': []})
    db = budget_service.db
    users = db.fetch_all(
        "SELECT u.id, u.nom, u.prenom FROM utilisateurs u "
        "WHERE u.service_id = %s AND u.actif = true ORDER BY u.nom, u.prenom",
        [service_id]
    )
    result = []
    for u in (users or []):
        budgets = db.fetch_all(
            "SELECT ba.id, ba.exercice, ba.nature, ba.statut, "
            "ba.montant_vote, ba.montant_engage, ba.montant_solde, bp.role as perm_role, "
            "e.code as entite_code, e.nom as entite_nom "
            "FROM budget_permissions bp "
            "JOIN budgets_annuels ba ON ba.id = bp.budget_id "
            "LEFT JOIN entites e ON e.id = ba.entite_id "
            "WHERE bp.user_id = %s ORDER BY ba.exercice DESC, e.code",
            [u['id']]
        )
        result.append({
            'user': dict(u),
            'budgets': [dict(b) for b in (budgets or [])]
        })
    return jsonify({'users': result})


@routes.route('/projets/vue_service', methods=['GET'])
@require_auth('gestionnaire_service')
def get_projets_vue_service():
    """Projets de chaque membre du service (gestionnaire_service uniquement)."""
    service_id = g.user.get('service_id')
    if not service_id:
        return jsonify({'users': []})
    db = budget_service.db
    users = db.fetch_all(
        "SELECT u.id, u.nom, u.prenom FROM utilisateurs u "
        "WHERE u.service_id = %s AND u.actif = true ORDER BY u.nom, u.prenom",
        [service_id]
    )
    result = []
    for u in (users or []):
        projets = db.fetch_all(
            "SELECT p.id, p.code, p.nom, p.statut, p.phase, p.type_projet, "
            "p.priorite, p.avancement, p.date_debut, p.date_fin_prevue, "
            "p.budget_estime, p.statut_rag "
            "FROM projets p "
            "WHERE p.created_by_id = %s ORDER BY p.statut, p.nom",
            [u['id']]
        )
        result.append({
            'user': dict(u),
            'projets': [dict(p) for p in (projets or [])]
        })
    return jsonify({'users': result})


# ─────────────────────────────────────────────
# BONS DE COMMANDE  — filtre par propriétaire
# ─────────────────────────────────────────────

def _ownership_where(user_id, role, service_id, alias='bc'):
    """
    Retourne (clause WHERE, params) pour filtrer par propriétaire.
    - admin   → voit tout
    - autres  → voit uniquement les siens (created_by_id = user_id)
    Les enregistrements sans created_by_id (NULL = données historiques)
    sont visibles uniquement par admin.
    """
    p = alias + '.'
    if role == 'admin':
        return "1=1", []
    return f"{p}created_by_id = %s", [user_id]


def _audit(action, table_name, record_id=None, details=None):
    """Enregistre une action dans le journal d'audit. N'interrompt jamais l'opération principale."""
    try:
        user = getattr(g, 'user', {}) or {}
        det  = json.dumps(details, ensure_ascii=False) if details and not isinstance(details, str) else details
        bc_service.db.execute(
            "INSERT INTO audit_log (user_id, user_login, action, table_name, record_id, details) "
            "VALUES (%s, %s, %s, %s, %s, %s)",
            [user.get('sub'), user.get('login'), action, table_name, record_id, det]
        )
    except Exception:
        pass


@routes.route('/bon_commande', methods=['GET'])
@require_auth()
def get_bon_commande():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    filters    = {k: v for k, v in request.args.items() if v}

    if role == 'admin':
        bc_list = bc_service.get_all_bons_commande(filters)
    else:
        # Récupérer tous les BCs accessibles puis appliquer filtres supplémentaires
        where, params = _ownership_where(user_id, role, service_id, 'bc')
        extra_where = [where]
        extra_params = list(params)
        if filters.get('statut'):
            extra_where.append("bc.statut = %s")
            extra_params.append(filters['statut'])
        if filters.get('entite_id'):
            extra_where.append("bc.entite_id = %s")
            extra_params.append(filters['entite_id'])
        rows = bc_service.db.fetch_all(
            "SELECT bc.*, f.nom as fournisseur_nom, e.code as entite_code, "
            "e.nom as entite_nom, p.nom as projet_nom, c.numero_contrat, "
            "u.nom || ' ' || COALESCE(u.prenom,'') as createur_nom "
            "FROM bons_commande bc "
            "LEFT JOIN fournisseurs f ON f.id = bc.fournisseur_id "
            "LEFT JOIN entites e ON e.id = bc.entite_id "
            "LEFT JOIN projets p ON p.id = bc.projet_id "
            "LEFT JOIN contrats c ON c.id = bc.contrat_id "
            "LEFT JOIN utilisateurs u ON u.id = bc.created_by_id "
            f"WHERE {' AND '.join(extra_where)} ORDER BY bc.date_creation DESC",
            extra_params
        )
        bc_list = [dict(r) for r in (rows or [])]
    return jsonify({"count": len(bc_list), "list": bc_list})


@routes.route('/bon_commande/stats', methods=['GET'])
@require_auth()
def get_bc_stats():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role == 'admin':
        return jsonify(bc_service.get_stats())
    where, params = _ownership_where(user_id, role, service_id, 'bc')
    row = bc_service.db.fetch_one(
        f"SELECT COUNT(*) as total, "
        f"SUM(CASE WHEN bc.statut='BROUILLON' THEN 1 ELSE 0 END) as brouillon, "
        f"SUM(CASE WHEN bc.statut='VALIDE' THEN 1 ELSE 0 END) as valide, "
        f"SUM(CASE WHEN bc.statut='SOLDE' THEN 1 ELSE 0 END) as solde, "
        f"COALESCE(SUM(bc.montant_ht),0) as total_ht "
        f"FROM bons_commande bc WHERE {where}",
        params
    )
    return jsonify(dict(row) if row else {})


@routes.route('/bon_commande/<int:bc_id>', methods=['GET'])
@require_auth()
def get_bon_commande_by_id(bc_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    bc = bc_service.get_by_id(bc_id)
    if not bc:
        return jsonify({"error": "BC introuvable"}), 404
    if role != 'admin':
        creator = bc.get('created_by_id')
        if creator is not None:
            where, params = _ownership_where(user_id, role, service_id, 'bc')
            row = bc_service.db.fetch_one(
                f"SELECT id FROM bons_commande bc WHERE bc.id=%s AND {where}",
                [bc_id] + params
            )
            if not row:
                return jsonify({"error": "Accès interdit"}), 403
    return jsonify(bc)


@routes.route('/bon_commande/<int:bc_id>/valider', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def valider_bc(bc_id):
    try:
        new_statut = bc_service.valider(bc_id)
        _audit('VALIDER', 'bons_commande', bc_id, {'nouveau_statut': new_statut})
        return jsonify({"success": True, "statut": new_statut})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/bon_commande/<int:bc_id>/imputer', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def imputer_bc(bc_id):
    data = request.json
    ligne_id = data.get('ligne_id')
    if not ligne_id:
        return jsonify({"success": False, "error": "ligne_id requis"}), 400
    try:
        bc_service.imputer(bc_id, ligne_id)
        _audit('IMPUTER', 'bons_commande', bc_id, {'ligne_id': ligne_id})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/bon_commande/<int:bc_id>/refuser', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def refuser_bon_commande(bc_id):
    data   = request.json or {}
    motif  = (data.get('motif') or '').strip()
    bc     = bc_service.get_by_id(bc_id)
    if not bc:
        return jsonify({"error": "BC introuvable"}), 404
    if bc.get('statut') not in ('EN_ATTENTE', 'VALIDE'):
        return jsonify({"error": f"Impossible de refuser un BC en statut '{bc.get('statut')}'"}), 400
    bc_service.db.execute(
        "UPDATE bons_commande SET statut='REFUSE', motif_refus=%s, date_maj=NOW() WHERE id=%s",
        [motif, bc_id]
    )
    _audit('REFUSER', 'bons_commande', bc_id, {'motif': motif} if motif else None)
    return jsonify({"success": True})


@routes.route('/bon_commande/parse_pdf', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def parse_bc_pdf():
    import io, re
    from difflib import get_close_matches
    try:
        import pdfplumber
    except ImportError:
        return jsonify({"error": "pdfplumber non installé — redéployez l'application"}), 500

    if 'file' not in request.files:
        return jsonify({"error": "Fichier PDF manquant"}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({"error": "Format PDF requis (.pdf)"}), 400

    try:
        text = ""
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"

        if not text.strip():
            return jsonify({"error": "Impossible d'extraire le texte du PDF (document scanné ?)"}), 422

        def _montant(val):
            if not val: return None
            try: return round(float(str(val).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')), 2)
            except: return None

        result = {}

        # ── N° BC ──────────────────────────────────────────────
        for p in [
            r'(?:bon\s+de\s+commande|commande\s+n[o°]?)[^\w]*([\w][\w\s\-\/]{1,30})',
            r'n[o°°]\s*(?:bc|commande|de\s+commande)\s*:?\s*([\w][\w\s\-\/]{1,30})',
            r'\bBC[\s\-]*([\d][\w\-\/]{1,20})',
        ]:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                result['numero_bc'] = m.group(1).strip()[:50]
                break

        # ── Objet ───────────────────────────────────────────────
        for p in [
            r'(?:objet|d[ée]signation|libell[ée]|prestation)\s*:?\s*(.{5,200}?)(?:\n|$)',
        ]:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                result['objet'] = m.group(1).strip()[:200]
                break

        # ── Montant TTC ─────────────────────────────────────────
        for p in [
            r'(?:total|montant)?\s*TTC\s*[:\s]+([\d][\d\s\.,]+)',
            r'T\.T\.C\.?\s*[:\s]+([\d][\d\s\.,]+)',
            r'([\d][\d\s\.,]+)\s*(?:€|EUR)\s*TTC',
        ]:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                v = _montant(m.group(1))
                if v and v > 0: result['montant_ttc'] = v; break

        # ── Montant HT ──────────────────────────────────────────
        for p in [
            r'(?:total|montant)?\s*HT\s*[:\s]+([\d][\d\s\.,]+)',
            r'H\.T\.?\s*[:\s]+([\d][\d\s\.,]+)',
            r'([\d][\d\s\.,]+)\s*(?:€|EUR)\s*HT',
        ]:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                v = _montant(m.group(1))
                if v and v > 0: result['montant_ht'] = v; break

        # ── Taux TVA ────────────────────────────────────────────
        m = re.search(r'TVA\s*:?\s*(\d+(?:[,\.]\d+)?)\s*%', text, re.IGNORECASE)
        if m:
            v = _montant(m.group(1))
            if v: result['tva'] = v

        # ── Déduire HT depuis TTC si manquant ───────────────────
        if result.get('montant_ttc') and not result.get('montant_ht'):
            tva = result.get('tva', 20)
            result['montant_ht'] = round(result['montant_ttc'] / (1 + tva / 100), 2)
        elif result.get('montant_ht') and not result.get('montant_ttc'):
            tva = result.get('tva', 20)
            result['montant_ttc'] = round(result['montant_ht'] * (1 + tva / 100), 2)

        # ── Fournisseur : fuzzy match ────────────────────────────
        fournisseurs = bc_service.db.fetch_all(
            "SELECT id, nom FROM fournisseurs ORDER BY nom"
        ) or []
        noms_fourn = [f['nom'] for f in fournisseurs]

        fourn_brut = None
        # 1) Chercher mot-clé "fournisseur:"
        for p in [r'(?:fournisseur|vendeur|prestataire|[ée]metteur)\s*:?\s*(.{3,80}?)(?:\n|$)']:
            m = re.search(p, text, re.IGNORECASE)
            if m: fourn_brut = m.group(1).strip(); break
        # 2) Chercher les 10 premières lignes non vides du PDF
        if not fourn_brut:
            lines = [l.strip() for l in text.split('\n') if l.strip() and len(l.strip()) > 3][:15]
            best_score = 0
            for line in lines:
                matches = get_close_matches(line, noms_fourn, n=1, cutoff=0.55)
                if matches:
                    fourn_brut = line
                    break

        result['fournisseur_nom_brut'] = fourn_brut
        if fourn_brut:
            matches = get_close_matches(fourn_brut, noms_fourn, n=1, cutoff=0.5)
            if matches:
                fourn = next((f for f in fournisseurs if f['nom'] == matches[0]), None)
                if fourn:
                    result['fournisseur_id']  = fourn['id']
                    result['fournisseur_nom'] = fourn['nom']

        # ── Suggestion ligne budgétaire (historique fournisseur) ─
        if result.get('fournisseur_id'):
            hist = bc_service.db.fetch_one(
                "SELECT lb.id, lb.libelle, COUNT(*) as cnt "
                "FROM bons_commande bc "
                "JOIN lignes_budgetaires lb ON lb.id = bc.ligne_budgetaire_id "
                "WHERE bc.fournisseur_id = %s AND bc.ligne_budgetaire_id IS NOT NULL "
                "GROUP BY lb.id, lb.libelle ORDER BY cnt DESC LIMIT 1",
                [result['fournisseur_id']]
            )
            if hist:
                result['ligne_budgetaire_id'] = hist['id']
                result['ligne_libelle']       = hist['libelle']

        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@routes.route('/bon_commande', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_bon_commande():
    data    = request.json
    user_id = g.user.get('sub')
    try:
        montant_ht  = float(data.get('montant_ht') or 0)
        montant_ttc = float(data.get('montant_ttc') or montant_ht * 1.2)
        bc_service.db.execute(
            "INSERT INTO bons_commande (numero_bc, objet, fournisseur_id, entite_id, "
            "projet_id, ligne_budgetaire_id, contrat_id, montant_ht, montant_ttc, statut, created_by_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            [data.get('numero_bc'), data.get('objet'), data.get('fournisseur_id') or None,
             data.get('entite_id') or None, data.get('projet_id') or None,
             data.get('ligne_budgetaire_id') or None, data.get('contrat_id') or None,
             montant_ht, montant_ttc, data.get('statut', 'BROUILLON'), user_id]
        )
        _audit('CREATE', 'bons_commande', None, {'numero_bc': data.get('numero_bc'), 'objet': data.get('objet'), 'montant_ttc': montant_ttc})
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/bon_commande/<int:bc_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_bon_commande(bc_id):
    data       = request.json
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'bc')
        row = bc_service.db.fetch_one(
            f"SELECT id FROM bons_commande bc WHERE bc.id=%s AND {where}", [bc_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    try:
        montant_ht  = float(data.get('montant_ht') or 0)
        montant_ttc = float(data.get('montant_ttc') or montant_ht * 1.2)
        new_statut   = data.get('statut')
        new_ligne_id = data.get('ligne_budgetaire_id') or None

        # Lire l'état AVANT la mise à jour pour gérer l'imputation comptable
        old_bc = bc_service.db.fetch_one(
            "SELECT statut, ligne_budgetaire_id, montant_ttc FROM bons_commande WHERE id=%s", [bc_id]
        )
        old_statut   = old_bc['statut'] if old_bc else None
        old_ligne_id = old_bc['ligne_budgetaire_id'] if old_bc else None
        old_montant  = float(old_bc['montant_ttc'] or 0) if old_bc else 0

        bc_service.db.execute(
            "UPDATE bons_commande SET numero_bc=%s, objet=%s, fournisseur_id=%s, "
            "entite_id=%s, projet_id=%s, ligne_budgetaire_id=%s, contrat_id=%s, "
            "montant_ht=%s, montant_ttc=%s, statut=%s, date_maj=NOW() WHERE id=%s",
            [data.get('numero_bc'), data.get('objet'), data.get('fournisseur_id') or None,
             data.get('entite_id') or None, data.get('projet_id') or None,
             new_ligne_id, data.get('contrat_id') or None,
             montant_ht, montant_ttc, new_statut, bc_id]
        )

        # Gestion comptable — recalcul depuis zéro après mise à jour du BC
        # (évite toute dérive liée aux deltas successifs)
        STATUTS_ENGAGES = ('VALIDE', 'IMPUTE', 'SOLDE')
        new_engage = new_statut in STATUTS_ENGAGES

        # Mettre à jour les flags d'imputation sur le BC lui-même
        if new_engage and new_ligne_id:
            bc_service.db.execute(
                "UPDATE bons_commande SET montant_engage=%s, "
                "date_imputation=COALESCE(date_imputation,NOW()), "
                "budget_impute=true, impute=true WHERE id=%s",
                [montant_ttc, bc_id]
            )
        elif not new_engage:
            bc_service.db.execute(
                "UPDATE bons_commande SET montant_engage=0, "
                "budget_impute=false, impute=false WHERE id=%s",
                [bc_id]
            )

        # Recalculer montant_engage de toutes les lignes affectées
        # (somme de tous les BCs engagés sur chaque ligne concernée)
        lignes_a_recalculer = set()
        if new_ligne_id:
            lignes_a_recalculer.add(int(new_ligne_id))
        if old_ligne_id:
            lignes_a_recalculer.add(int(old_ligne_id))

        for lid in lignes_a_recalculer:
            bc_service.db.execute(
                "UPDATE lignes_budgetaires l "
                "SET montant_engage = ("
                "  SELECT COALESCE(SUM(bc.montant_ttc), 0) "
                "  FROM bons_commande bc "
                "  WHERE bc.ligne_budgetaire_id = l.id "
                "  AND bc.statut IN ('VALIDE', 'IMPUTE', 'SOLDE')"
                "), "
                "montant_solde = montant_vote - ("
                "  SELECT COALESCE(SUM(bc.montant_ttc), 0) "
                "  FROM bons_commande bc "
                "  WHERE bc.ligne_budgetaire_id = l.id "
                "  AND bc.statut IN ('VALIDE', 'IMPUTE', 'SOLDE')"
                "), "
                "date_maj=NOW() "
                "WHERE l.id = %s",
                [lid]
            )

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/bon_commande/<int:bc_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def delete_bon_commande(bc_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'bc')
        row = bc_service.db.fetch_one(
            f"SELECT id FROM bons_commande bc WHERE bc.id=%s AND {where}", [bc_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    try:
        bc_service.db.execute("DELETE FROM bons_commande WHERE id=%s", [bc_id])
        _audit('DELETE', 'bons_commande', bc_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─────────────────────────────────────────────
# CONTRATS — filtre par propriétaire
# ─────────────────────────────────────────────

@routes.route('/contrat', methods=['GET'])
@require_auth()
def get_contrats():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role == 'admin':
        contrats = contrat_service.get_all()
    else:
        where, params = _ownership_where(user_id, role, service_id, 'c')
        rows = contrat_service.db.fetch_all(
            "SELECT c.*, f.nom as fournisseur_nom, "
            "u.nom || ' ' || COALESCE(u.prenom,'') as createur_nom "
            "FROM contrats c "
            "LEFT JOIN fournisseurs f ON f.id = c.fournisseur_id "
            "LEFT JOIN utilisateurs u ON u.id = c.created_by_id "
            f"WHERE {where} ORDER BY c.date_creation DESC",
            params
        )
        contrats = [dict(r) for r in (rows or [])]
    return jsonify({"count": len(contrats), "list": contrats})


@routes.route('/contrat/<int:contrat_id>', methods=['GET'])
@require_auth()
def get_contrat_by_id(contrat_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    c = contrat_service.get_by_id(contrat_id)
    if not c:
        return jsonify({"error": "Contrat introuvable"}), 404
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'c')
        row = contrat_service.db.fetch_one(
            f"SELECT id FROM contrats c WHERE c.id=%s AND {where}", [contrat_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    return jsonify(c)


@routes.route('/contrat/alertes', methods=['GET'])
@require_auth()
def get_contrat_alertes():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role == 'admin':
        return jsonify({"list": contrat_service.get_alertes()})
    where, params = _ownership_where(user_id, role, service_id, 'c')
    rows = contrat_service.db.fetch_all(
        "SELECT c.*, f.nom as fournisseur_nom FROM contrats c "
        "LEFT JOIN fournisseurs f ON f.id = c.fournisseur_id "
        f"WHERE c.date_fin IS NOT NULL AND c.date_fin <= NOW() + INTERVAL '60 days' AND {where} "
        "ORDER BY c.date_fin",
        params
    )
    return jsonify({"list": [dict(r) for r in (rows or [])]})


@routes.route('/contrat/<int:contrat_id>/reconduire', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def reconduire_contrat(contrat_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'c')
        row = contrat_service.db.fetch_one(
            f"SELECT id FROM contrats c WHERE c.id=%s AND {where}", [contrat_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    data = request.json
    nouvelle_date_fin = data.get('nouvelle_date_fin')
    if not nouvelle_date_fin:
        return jsonify({"success": False, "error": "nouvelle_date_fin requise"}), 400
    try:
        contrat_service.reconduire(contrat_id, nouvelle_date_fin)
        _audit('RECONDUIRE', 'contrats', contrat_id, {'nouvelle_date_fin': nouvelle_date_fin})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/contrat', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_contrat():
    data    = request.json
    user_id = g.user.get('sub')
    try:
        montant_ht = float(data.get('montant_total_ht') or 0)
        contrat_service.db.execute(
            "INSERT INTO contrats (numero_contrat, objet, fournisseur_id, montant_initial_ht, "
            "montant_total_ht, montant_ttc, date_debut, date_fin, statut, created_by_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            [data.get('numero_contrat'), data.get('objet'), data.get('fournisseur_id') or None,
             montant_ht, montant_ht, round(montant_ht * 1.2, 2),
             data.get('date_debut') or None, data.get('date_fin') or None,
             data.get('statut', 'ACTIF'), user_id]
        )
        _audit('CREATE', 'contrats', None, {'numero': data.get('numero_contrat'), 'objet': data.get('objet')})
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/contrat/<int:contrat_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_contrat(contrat_id):
    data       = request.json
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'c')
        row = contrat_service.db.fetch_one(
            f"SELECT id FROM contrats c WHERE c.id=%s AND {where}", [contrat_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    try:
        montant_ht = float(data.get('montant_total_ht') or 0)
        contrat_service.db.execute(
            "UPDATE contrats SET numero_contrat=%s, objet=%s, fournisseur_id=%s, "
            "montant_total_ht=%s, montant_ttc=%s, date_debut=%s, date_fin=%s, "
            "statut=%s, type_marche=%s, date_maj=NOW() WHERE id=%s",
            [data.get('numero_contrat'), data.get('objet'), data.get('fournisseur_id') or None,
             montant_ht, round(montant_ht * 1.2, 2),
             data.get('date_debut') or None, data.get('date_fin') or None,
             data.get('statut'), data.get('type_marche') or None, contrat_id]
        )
        return _ok()
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/contrat/<int:contrat_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def delete_contrat(contrat_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'c')
        row = contrat_service.db.fetch_one(
            f"SELECT id FROM contrats c WHERE c.id=%s AND {where}", [contrat_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    try:
        contrat_service.db.execute("DELETE FROM contrats WHERE id=%s", [contrat_id])
        _audit('DELETE', 'contrats', contrat_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─────────────────────────────────────────────
# PROJETS
# ─────────────────────────────────────────────

@routes.route('/projet', methods=['GET'])
@require_auth()
def get_projets():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    filters    = request.args.to_dict()

    if role == 'admin':
        projets = projet_service.get_all(filters)
    else:
        where, params = _ownership_where(user_id, role, service_id, 'p')
        extra = []
        if filters.get('statut'):
            extra.append("p.statut = %s")
            params.append(filters['statut'])
        clause = where + (" AND " + " AND ".join(extra) if extra else "")
        rows = projet_service.db.fetch_all(
            "SELECT p.*, s.nom as service_nom, s.code as service_code "
            "FROM projets p "
            "LEFT JOIN services s ON s.id = p.service_id "
            f"WHERE {clause} ORDER BY p.date_creation DESC",
            params
        )
        projets = [dict(r) for r in (rows or [])]
    return jsonify({"count": len(projets), "list": projets})


@routes.route('/projet/<int:projet_id>', methods=['GET'])
@require_auth()
def get_projet(projet_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    p = projet_service.get_by_id(projet_id)
    if not p:
        return jsonify({"error": "Projet introuvable"}), 404
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'p')
        row = projet_service.db.fetch_one(
            f"SELECT id FROM projets p WHERE p.id=%s AND {where}", [projet_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403
    return jsonify(p)

_PROJET_RULES = {
    'code': {'label': 'Code projet', 'required': True, 'max': 50},
    'nom':  {'label': 'Intitulé',    'required': True, 'max': 300},
    'statut': {'label': 'Statut', 'enum': ['ACTIF','EN_ATTENTE','TERMINE','ANNULE','EN_PAUSE']},
    'priorite': {'label': 'Priorité', 'enum': ['CRITIQUE','HAUTE','MOYENNE','BASSE','']},
    'avancement': {'label': 'Avancement', 'type': 'number'},
}

@routes.route('/projet', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_projet():
    data    = request.json or {}
    user_id = g.user.get('sub')
    errs = _validate(data, _PROJET_RULES)
    if errs:
        return _err(' | '.join(errs))
    try:
        projet_service.db.execute(
            "INSERT INTO projets (code, nom, description, statut, priorite, type_projet, phase, "
            "service_id, date_debut, date_fin_prevue, date_fin_reelle, "
            "budget_initial, budget_estime, budget_actuel, avancement, "
            "objectifs, enjeux, gains, risques, contraintes, solutions, created_by_id) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            [data.get('code'), data.get('nom'), data.get('description'), data.get('statut'),
             data.get('priorite'), data.get('type_projet') or None, data.get('phase') or None,
             data.get('service_id') or None,
             data.get('date_debut') or None, data.get('date_fin_prevue') or None,
             data.get('date_fin_reelle') or None,
             data.get('budget_initial') or None, data.get('budget_estime') or None,
             data.get('budget_actuel') or None, data.get('avancement') or 0,
             data.get('objectifs') or None, data.get('enjeux') or None,
             data.get('gains') or None, data.get('risques') or None,
             data.get('contraintes') or None, data.get('solutions') or None, user_id]
        )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/projet/<int:projet_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_projet(projet_id):
    data = request.json or {}
    errs = _validate(data, _PROJET_RULES)
    if errs:
        return _err(' | '.join(errs))
    try:
        import json as _json
        def _jstr(v):
            if v is None: return None
            return v if isinstance(v, str) else _json.dumps(v, ensure_ascii=False)
        projet_service.db.execute(
            "UPDATE projets SET code=%s, nom=%s, description=%s, statut=%s, priorite=%s, "
            "type_projet=%s, phase=%s, service_id=%s, "
            "date_debut=%s, date_fin_prevue=%s, date_fin_reelle=%s, "
            "budget_initial=%s, budget_estime=%s, budget_actuel=%s, avancement=%s, "
            "objectifs=%s, enjeux=%s, gains=%s, risques=%s, contraintes=%s, solutions=%s, "
            "financement=%s, registre_risques=%s, contraintes_6axes=%s, "
            "triangle_tensions=%s, arbitrage=%s, "
            "chef_projet_contact_id=%s, responsable_contact_id=%s, note=%s, "
            "updated_at=NOW() WHERE id=%s",
            [data.get('code'), data.get('nom'), data.get('description'), data.get('statut'),
             data.get('priorite'), data.get('type_projet') or None, data.get('phase') or None,
             data.get('service_id') or None,
             data.get('date_debut') or None, data.get('date_fin_prevue') or None,
             data.get('date_fin_reelle') or None,
             data.get('budget_initial') or None, data.get('budget_estime') or None,
             data.get('budget_actuel') or None, data.get('avancement') or 0,
             data.get('objectifs') or None, data.get('enjeux') or None,
             data.get('gains') or None, data.get('risques') or None,
             data.get('contraintes') or None, data.get('solutions') or None,
             data.get('financement') or None,
             _jstr(data.get('registre_risques')) or None,
             _jstr(data.get('contraintes_6axes')) or None,
             _jstr(data.get('triangle_tensions')) or None,
             data.get('arbitrage') or None,
             data.get('chef_projet_contact_id') or None,
             data.get('responsable_contact_id') or None,
             data.get('note') or None,
             projet_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/projet/<int:projet_id>', methods=['DELETE'])
@require_auth('admin')
def delete_projet(projet_id):
    try:
        projet_service.db.execute("DELETE FROM projets WHERE id=%s", [projet_id])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/projet/<int:projet_id>/fiche_word', methods=['GET'])
@require_auth()
def export_fiche_projet_word(projet_id):
    """Génère et télécharge la fiche projet en .docx"""
    import tempfile
    from flask import send_file
    from app.services.fiche_projet_web_service import generer_fiche_depuis_id_pg

    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')

    p = projet_service.get_by_id(projet_id)
    if not p:
        return jsonify({"error": "Projet introuvable"}), 404
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'p')
        row = projet_service.db.fetch_one(
            f"SELECT id FROM projets p WHERE p.id=%s AND {where}", [projet_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403

    try:
        tmpdir = tempfile.mkdtemp()
        out_path = generer_fiche_depuis_id_pg(projet_id, tmpdir, projet_service.db)
        code = p.get('code', f'PRJ{projet_id}')
        return send_file(
            out_path,
            as_attachment=True,
            download_name=f"fiche_projet_{code}.docx",
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@routes.route('/projet/<int:projet_id>/fiche_html', methods=['GET'])
@require_auth()
def export_fiche_projet_html(projet_id):
    """Génère et retourne la fiche projet en HTML."""
    from flask import make_response
    from app.services.fiche_projet_html_service import generer_fiche_html_depuis_id_pg

    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')

    p = projet_service.get_by_id(projet_id)
    if not p:
        return jsonify({"error": "Projet introuvable"}), 404
    if role != 'admin':
        where, params = _ownership_where(user_id, role, service_id, 'p')
        row = projet_service.db.fetch_one(
            f"SELECT id FROM projets p WHERE p.id=%s AND {where}", [projet_id] + params
        )
        if not row:
            return jsonify({"error": "Accès interdit"}), 403

    try:
        html_content = generer_fiche_html_depuis_id_pg(projet_id, projet_service.db)
        resp = make_response(html_content)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@routes.route('/projet/<int:projet_id>/equipe', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def add_projet_equipe(projet_id):
    data = request.json or {}
    try:
        projet_service.add_equipe_membre(
            projet_id,
            data.get('utilisateur_id') or None,
            data.get('membre_label') or None
        )
        return jsonify({"ok": True}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@routes.route('/projet/<int:projet_id>/equipe/<int:membre_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def remove_projet_equipe(projet_id, membre_id):
    try:
        projet_service.remove_equipe_membre(projet_id, membre_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@routes.route('/projet/<int:projet_id>/contact', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def add_projet_contact(projet_id):
    data = request.json or {}
    contact_id    = data.get('contact_id') or None
    contact_libre = (data.get('contact_libre') or '').strip() or None
    if not contact_id and not contact_libre:
        return jsonify({"error": "contact_id ou contact_libre requis"}), 400
    try:
        projet_service.add_projet_contact(projet_id, contact_id, data.get('role'), contact_libre)
        return jsonify({"ok": True}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@routes.route('/projet/<int:projet_id>/contact/<int:contact_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def remove_projet_contact(projet_id, contact_id):
    try:
        projet_service.remove_projet_contact(projet_id, contact_id=contact_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@routes.route('/projet/<int:projet_id>/contact/libre', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def remove_projet_contact_libre(projet_id):
    data = request.json or {}
    try:
        projet_service.remove_projet_contact(projet_id, contact_libre=data.get('contact_libre'))
        return _ok()
    except Exception as e:
        return _err(e)


# ─────────────────────────────────────────────
# JALONS
# ─────────────────────────────────────────────

@routes.route('/projet/<int:projet_id>/jalons', methods=['GET'])
@require_auth()
def get_jalons(projet_id):
    try:
        rows = projet_service.db.fetch_all(
            "SELECT * FROM jalons WHERE projet_id=%s ORDER BY date_echeance ASC NULLS LAST",
            [projet_id]
        )
        return jsonify({"list": [dict(r) for r in (rows or [])]})
    except Exception as e:
        return jsonify({"list": [], "warning": str(e)}), 200


@routes.route('/projet/<int:projet_id>/jalons', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def add_jalon(projet_id):
    data = request.json or {}
    user_id = g.user.get('sub')
    try:
        projet_service.db.execute(
            "INSERT INTO jalons (projet_id, titre, date_echeance, statut, description, created_by_id) "
            "VALUES (%s, %s, %s, %s, %s, %s)",
            [projet_id, data.get('titre'), data.get('date_echeance') or None,
             data.get('statut', 'A_VENIR'), data.get('description') or None, user_id]
        )
        return _ok(), 201
    except Exception as e:
        return _err(e)


@routes.route('/projet/<int:projet_id>/jalons/<int:jalon_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_jalon(projet_id, jalon_id):
    data = request.json or {}
    try:
        projet_service.db.execute(
            "UPDATE jalons SET titre=%s, date_echeance=%s, statut=%s, description=%s "
            "WHERE id=%s AND projet_id=%s",
            [data.get('titre'), data.get('date_echeance') or None,
             data.get('statut'), data.get('description') or None, jalon_id, projet_id]
        )
        return _ok()
    except Exception as e:
        return _err(e)


@routes.route('/projet/<int:projet_id>/jalons/<int:jalon_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def delete_jalon(projet_id, jalon_id):
    try:
        projet_service.db.execute(
            "DELETE FROM jalons WHERE id=%s AND projet_id=%s", [jalon_id, projet_id]
        )
        return _ok()
    except Exception as e:
        return _err(e)


# ─────────────────────────────────────────────
# JOURNAL DE BORD PROJET
# ─────────────────────────────────────────────

@routes.route('/projet/<int:projet_id>/journal', methods=['GET'])
@require_auth()
def get_journal_projet(projet_id):
    try:
        rows = projet_service.db.fetch_all(
            "SELECT * FROM journal_projet WHERE projet_id=%s ORDER BY date_entree DESC",
            [projet_id]
        )
        return jsonify({"list": [dict(r) for r in (rows or [])]})
    except Exception as e:
        return jsonify({"list": [], "warning": str(e)}), 200


@routes.route('/projet/<int:projet_id>/journal', methods=['POST'])
@require_auth()
def add_journal_entry(projet_id):
    data = request.json or {}
    user_id = g.user.get('sub')
    nom    = (g.user.get('prenom', '') + ' ' + g.user.get('nom', '')).strip()
    auteur = nom or g.user.get('login', 'Utilisateur')
    try:
        projet_service.db.execute(
            "INSERT INTO journal_projet (projet_id, auteur, type_entree, contenu, created_by_id) "
            "VALUES (%s, %s, %s, %s, %s)",
            [projet_id, auteur, data.get('type_entree', 'EVENEMENT'),
             data.get('contenu', ''), user_id]
        )
        return _ok(), 201
    except Exception as e:
        return _err(e)


@routes.route('/projet/<int:projet_id>/journal/<int:entry_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def delete_journal_entry(projet_id, entry_id):
    try:
        projet_service.db.execute(
            "DELETE FROM journal_projet WHERE id=%s AND projet_id=%s", [entry_id, projet_id]
        )
        return _ok()
    except Exception as e:
        return _err(e)


# ─────────────────────────────────────────────
# RAG PROJET
# ─────────────────────────────────────────────

@routes.route('/projet/<int:projet_id>/rag', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_rag(projet_id):
    data = request.json or {}
    rag = data.get('statut_rag', 'VERT')
    if rag not in ('ROUGE', 'AMBER', 'VERT'):
        return _err("Valeur RAG invalide")
    try:
        projet_service.db.execute(
            "UPDATE projets SET statut_rag=%s WHERE id=%s", [rag, projet_id]
        )
        return _ok()
    except Exception as e:
        return _err(e)


# ─────────────────────────────────────────────
# TÂCHES
# ─────────────────────────────────────────────

def _tache_visibility_where(user_id, role, service_id):
    """
    Filtre de visibilité pour les tâches :
    - admin        → voit tout
    - gestionnaire → voit toutes les tâches assignées à son service/unité (+ les siennes)
    - lecteur      → voit uniquement les tâches qui lui sont assignées ou qu'il a créées
    Les tâches sans assignee (NULL) sont visibles aux gestionnaires et admins.
    """
    if role == 'admin':
        return "1=1", []
    return "(t.assignee_id = %s OR t.created_by_id = %s)", [user_id, user_id]


@routes.route('/tache', methods=['GET'])
@require_auth()
def get_taches():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role == 'admin':
        taches = tache_service.get_all()
    else:
        where, params = _tache_visibility_where(user_id, role, service_id)
        rows = tache_service.db.fetch_all(
            "SELECT t.*, p.nom as projet_nom, p.code as projet_code, "
            "u.nom || ' ' || u.prenom as assignee_nom, "
            "u.id as assignee_user_id, "
            "s.nom as assignee_service_nom, s.code as assignee_service_code, "
            "s.is_unite as assignee_is_unite "
            "FROM taches t "
            "LEFT JOIN projets p ON p.id = t.projet_id "
            "LEFT JOIN utilisateurs u ON u.id = t.assignee_id "
            "LEFT JOIN services s ON s.id = u.service_id "
            f"WHERE {where} "
            "ORDER BY t.date_echeance ASC NULLS LAST, t.id DESC",
            params
        )
        taches = [dict(r) for r in (rows or [])]
    return jsonify({"count": len(taches), "list": taches})


@routes.route('/tache', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_tache():
    data    = request.json
    user_id = g.user.get('sub')
    try:
        tache_service.db.execute(
            "INSERT INTO taches (projet_id, titre, statut, priorite, date_debut, date_echeance, "
            "estimation_heures, avancement, responsable_label, assignee_id, created_by_id, "
            "type_tache, rapport_reunion) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            [data.get('projet_id') or None, data.get('titre'),
             data.get('statut', 'A faire'), data.get('priorite'),
             data.get('date_debut') or None, data.get('date_echeance') or None,
             data.get('estimation_heures') or None, data.get('avancement') or 0,
             data.get('responsable_label') or None,
             data.get('assignee_id') or None, user_id,
             data.get('type_tache', 'autre'), data.get('rapport_reunion') or None]
        )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/tache/<int:tache_id>', methods=['GET'])
@require_auth()
def get_tache_by_id(tache_id):
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    try:
        if role != 'admin':
            vis_w, vis_p = _tache_visibility_where(user_id, role, service_id)
            row = tache_service.db.fetch_one(
                "SELECT t.*, p.nom as projet_nom FROM taches t "
                "LEFT JOIN projets p ON p.id = t.projet_id "
                f"WHERE t.id = %s AND ({vis_w})",
                [tache_id] + vis_p
            )
        else:
            row = tache_service.db.fetch_one(
                "SELECT t.*, p.nom as projet_nom FROM taches t "
                "LEFT JOIN projets p ON p.id = t.projet_id WHERE t.id = %s",
                [tache_id]
            )
        if row:
            return jsonify(dict(row))
        return jsonify({"error": "Tâche introuvable"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@routes.route('/tache/<int:tache_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_tache(tache_id):
    data       = request.json
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    # Vérifier que la tâche est visible avant modification
    if role != 'admin':
        vis_w, vis_p = _tache_visibility_where(user_id, role, service_id)
        check = tache_service.db.fetch_one(
            f"SELECT id FROM taches t WHERE t.id=%s AND ({vis_w})",
            [tache_id] + vis_p
        )
        if not check:
            return jsonify({"error": "Tâche introuvable"}), 404
    try:
        tache_service.db.execute(
            "UPDATE taches SET projet_id=%s, titre=%s, statut=%s, priorite=%s, "
            "date_debut=%s, date_echeance=%s, estimation_heures=%s, avancement=%s, "
            "responsable_label=%s, assignee_id=%s, type_tache=%s, rapport_reunion=%s, "
            "updated_at=NOW() WHERE id=%s",
            [data.get('projet_id') or None, data.get('titre'), data.get('statut'),
             data.get('priorite'), data.get('date_debut') or None,
             data.get('date_echeance') or None,
             data.get('estimation_heures') or None, data.get('avancement') or 0,
             data.get('responsable_label') or None,
             data.get('assignee_id') or None,
             data.get('type_tache', 'autre'), data.get('rapport_reunion') or None,
             tache_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/tache/<int:tache_id>', methods=['DELETE'])
@require_auth('admin')
def delete_tache(tache_id):
    try:
        tache_service.db.execute("DELETE FROM taches WHERE id=%s", [tache_id])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─────────────────────────────────────────────
# FOURNISSEURS
# ─────────────────────────────────────────────

@routes.route('/fournisseur', methods=['GET'])
@require_auth()
def get_fournisseurs_list():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    try:
        if role == 'admin':
            w_clause, w_params = "1=1", []
        else:
            w_clause, w_params = _ownership_where(user_id, role, service_id, 'f')
        rows = referentiel_service.db.fetch_all(
            "SELECT f.*, "
            "(SELECT COUNT(*) FROM contrats c WHERE c.fournisseur_id=f.id) as nb_contrats, "
            "(SELECT COUNT(*) FROM bons_commande bc WHERE bc.fournisseur_id=f.id) as nb_bc, "
            "(SELECT COALESCE(SUM(bc.montant_ttc),0) FROM bons_commande bc WHERE bc.fournisseur_id=f.id) as montant_total, "
            "(SELECT STRING_AGG(TRIM(CONCAT(c.nom, ' ', COALESCE(c.prenom,''))), ', ' ORDER BY c.nom) "
            " FROM contacts c JOIN fournisseur_contacts fc ON fc.contact_id=c.id WHERE fc.fournisseur_id=f.id) as contacts_lies "
            f"FROM fournisseurs f WHERE {w_clause} ORDER BY f.nom",
            w_params
        )
        result = [dict(r) for r in rows] if rows else []
        return jsonify({"count": len(result), "list": result})
    except Exception as e:
        return jsonify({"count": 0, "list": [], "error": str(e)})


@routes.route('/fournisseur', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_fournisseur():
    data    = request.json
    user_id = g.user.get('sub')
    try:
        referentiel_service.db.execute(
            "INSERT INTO fournisseurs (nom, contact_principal, email, telephone, adresse, ville, statut, created_by_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            [data.get('nom'), data.get('contact_principal'), data.get('email'),
             data.get('telephone'), data.get('adresse'), data.get('ville'), 'ACTIF', user_id]
        )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/fournisseur/<int:fournisseur_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_fournisseur(fournisseur_id):
    data       = request.json
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    if role != 'admin':
        own_w, own_p = _ownership_where(user_id, role, service_id, 'f')
        check = referentiel_service.db.fetch_one(
            f"SELECT id FROM fournisseurs f WHERE f.id=%s AND {own_w}",
            [fournisseur_id] + own_p
        )
        if not check:
            return jsonify({"error": "Fournisseur introuvable"}), 404
    try:
        referentiel_service.db.execute(
            "UPDATE fournisseurs SET nom=%s, contact_principal=%s, email=%s, "
            "telephone=%s, adresse=%s, ville=%s WHERE id=%s",
            [data.get('nom'), data.get('contact_principal'), data.get('email'),
             data.get('telephone'), data.get('adresse'), data.get('ville'), fournisseur_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/fournisseur/<int:fournisseur_id>', methods=['DELETE'])
@require_auth('admin')
def delete_fournisseur(fournisseur_id):
    try:
        referentiel_service.db.execute("DELETE FROM fournisseurs WHERE id=%s", [fournisseur_id])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ── Contacts liés à un fournisseur ──────────────────────────

@routes.route('/fournisseur/<int:fournisseur_id>/contacts', methods=['GET'])
@require_auth()
def get_fournisseur_contacts(fournisseur_id):
    try:
        rows = referentiel_service.db.fetch_all(
            "SELECT c.id, c.nom, c.prenom, c.fonction, c.type, c.telephone, c.email, c.societe "
            "FROM contacts c "
            "JOIN fournisseur_contacts fc ON fc.contact_id = c.id "
            "WHERE fc.fournisseur_id = %s ORDER BY c.nom, c.prenom",
            [fournisseur_id]
        )
        return jsonify({"list": [dict(r) for r in rows] if rows else []})
    except Exception as e:
        return jsonify({"list": [], "error": str(e)})


@routes.route('/fournisseur/<int:fournisseur_id>/contacts', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def add_fournisseur_contact(fournisseur_id):
    data = request.json
    contact_id = data.get('contact_id')
    if not contact_id:
        return jsonify({"error": "contact_id requis"}), 400
    try:
        referentiel_service.db.execute(
            "INSERT INTO fournisseur_contacts (fournisseur_id, contact_id) "
            "VALUES (%s, %s) ON CONFLICT DO NOTHING",
            [fournisseur_id, contact_id]
        )
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/fournisseur/<int:fournisseur_id>/contacts/<int:contact_id>', methods=['DELETE'])
@require_auth('admin', 'gestionnaire')
def remove_fournisseur_contact(fournisseur_id, contact_id):
    try:
        referentiel_service.db.execute(
            "DELETE FROM fournisseur_contacts WHERE fournisseur_id=%s AND contact_id=%s",
            [fournisseur_id, contact_id]
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─────────────────────────────────────────────
# CONTACTS
# ─────────────────────────────────────────────

@routes.route('/contact', methods=['GET'])
@require_auth()
def get_contacts():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    filters    = request.args.to_dict()

    if role == 'admin':
        contacts = contact_service.get_all(filters)
    else:
        where, params = _ownership_where(user_id, role, service_id, 'c')
        extra = []
        if filters.get('type'):
            extra.append("AND c.type = %s")
            params.append(filters['type'])
        if filters.get('search'):
            s = '%' + filters['search'] + '%'
            extra.append("AND (c.nom ILIKE %s OR c.prenom ILIKE %s OR c.email ILIKE %s OR c.organisation ILIKE %s)")
            params.extend([s, s, s, s])
        rows = contact_service.db.fetch_all(
            "SELECT c.*, s.nom as service_nom "
            "FROM contacts c "
            "LEFT JOIN services s ON s.id = c.service_id "
            f"WHERE {where} " + " ".join(extra) +
            " ORDER BY c.nom, c.prenom",
            params
        )
        contacts = [dict(r) for r in rows] if rows else []

    return jsonify({"count": len(contacts), "list": contacts})

_CONTACT_RULES = {
    'nom':   {'label': 'Nom',   'required': True, 'max': 200},
    'prenom':{'label': 'Prénom','max': 200},
    'email': {'label': 'Email', 'type': 'email', 'max': 200},
    'type':  {'label': 'Type',  'enum': ['INTERNE','EXTERNE','PRESTATAIRE','PARTENAIRE','DIRECTION','SERVICE','']},
}

@routes.route('/contact', methods=['POST'])
@require_auth('admin', 'gestionnaire')
def create_contact():
    data = request.json or {}
    errs = _validate(data, _CONTACT_RULES)
    if errs:
        return _err(' | '.join(errs))
    data['created_by_id'] = g.user.get('sub')
    try:
        contact_service.create(data)
        return _ok(), 201
    except Exception as e:
        return _err(e)

@routes.route('/contact/<int:contact_id>', methods=['PUT'])
@require_auth('admin', 'gestionnaire')
def update_contact(contact_id):
    data = request.json or {}
    errs = _validate(data, _CONTACT_RULES)
    if errs:
        return _err(' | '.join(errs))
    try:
        contact_service.update(contact_id, data)
        return _ok()
    except Exception as e:
        return _err(e)

@routes.route('/contact/<int:contact_id>', methods=['DELETE'])
@require_auth('admin')
def delete_contact(contact_id):
    try:
        contact_service.delete(contact_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


# ─────────────────────────────────────────────
# SERVICES (ORGANISATION)
# ─────────────────────────────────────────────

@routes.route('/service_org', methods=['GET'])
@require_auth()
def get_services_org():
    services = service_org_service.get_all()
    return jsonify({"count": len(services), "list": services})

@routes.route('/service_org', methods=['POST'])
@require_auth('admin')
def create_service_org():
    data = request.json
    try:
        service_org_service.create(data)
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/service_org/<int:service_id>', methods=['PUT'])
@require_auth('admin')
def update_service_org(service_id):
    data = request.json
    try:
        service_org_service.update(service_id, data)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/service_org/<int:service_id>', methods=['DELETE'])
@require_auth('admin')
def delete_service_org(service_id):
    try:
        # Vérifier les dépendances avant suppression
        db = service_org_service.db
        nb_projets = db.fetch_one(
            "SELECT COUNT(*) AS n FROM projets WHERE service_id = %s", [service_id]
        )
        nb_users = db.fetch_one(
            "SELECT COUNT(*) AS n FROM utilisateurs WHERE service_id = %s AND actif = true", [service_id]
        )
        nb_enfants = db.fetch_one(
            "SELECT COUNT(*) AS n FROM services WHERE parent_id = %s", [service_id]
        )
        msgs = []
        if nb_projets and int(nb_projets['n']) > 0:
            msgs.append(f"{nb_projets['n']} projet(s) lié(s)")
        if nb_users and int(nb_users['n']) > 0:
            msgs.append(f"{nb_users['n']} utilisateur(s) rattaché(s)")
        if nb_enfants and int(nb_enfants['n']) > 0:
            msgs.append(f"{nb_enfants['n']} service(s)/unité(s) enfant(s)")
        if msgs:
            return jsonify({
                "success": False,
                "error": f"Impossible de supprimer : {', '.join(msgs)}. Réattribuez-les d'abord."
            }), 400
        service_org_service.delete(service_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

@routes.route('/service_org/<int:service_id>/membres', methods=['GET'])
@require_auth()
def get_service_membres(service_id):
    membres = service_org_service.get_membres(service_id)
    return jsonify({"list": membres})


# ─────────────────────────────────────────────
# RÉFÉRENTIELS (données pour les selects)
# ─────────────────────────────────────────────

@routes.route('/referentiels', methods=['GET'])
@require_auth()
def get_referentiels():
    try:
        etp          = referentiel_service.get_etp()
        fournisseurs = referentiel_service.get_fournisseurs()
        contacts     = referentiel_service.get_contacts()
        services     = referentiel_service.get_services()
        entites      = budget_service.get_entites()
        projets      = projet_service.get_all()
        lignes       = budget_service.get_lignes()
        applications = budget_service.get_all_applications()
        contrats_ref = contrat_service.db.fetch_all(
            "SELECT id, numero_contrat, objet FROM contrats ORDER BY numero_contrat"
        )
        contrats_ref = [dict(c) for c in contrats_ref] if contrats_ref else []
        return jsonify({
            "etp": etp, "fournisseurs": fournisseurs,
            "contacts": contacts, "services": services,
            "entites": entites, "projets": projets,
            "lignes": lignes, "applications": applications,
            "contrats": contrats_ref,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# KANBAN
# ─────────────────────────────────────────────

@routes.route('/kanban', methods=['GET'])
@require_auth()
def kanban():
    cur_user_id = g.user.get('sub')
    role        = g.user.get('role')
    service_id  = g.user.get('service_id')
    projet_id   = request.args.get('projet_id')
    user_id     = request.args.get('user_id')

    if role == 'admin':
        taches = tache_service.get_all()
    else:
        where, params = _tache_visibility_where(cur_user_id, role, service_id)
        rows = tache_service.db.fetch_all(
            "SELECT t.*, p.nom as projet_nom, p.code as projet_code, "
            "u.nom || ' ' || u.prenom as assignee_nom, u.id as assignee_user_id, "
            "s.nom as assignee_service_nom, s.code as assignee_service_code, "
            "s.is_unite as assignee_is_unite "
            "FROM taches t "
            "LEFT JOIN projets p ON p.id = t.projet_id "
            "LEFT JOIN utilisateurs u ON u.id = t.assignee_id "
            "LEFT JOIN services s ON s.id = u.service_id "
            f"WHERE {where} ORDER BY t.date_echeance ASC NULLS LAST",
            params
        )
        taches = [dict(r) for r in (rows or [])]

    if projet_id:
        taches = [t for t in taches if str(t.get('projet_id', '')) == str(projet_id)]
    if user_id:
        taches = [t for t in taches if str(t.get('assignee_id', '')) == str(user_id)]
    colonnes = ['A faire', 'En cours', 'En attente', 'Bloqué', 'Terminé']
    columns = {c: [] for c in colonnes}
    for t in taches:
        statut = t.get('statut', 'A faire')
        if statut in columns:
            columns[statut].append(t)
        else:
            columns.setdefault(statut, []).append(t)
    return jsonify({"columns": columns})


# ─────────────────────────────────────────────
# ETP / CHARGE (simplifié)
# ─────────────────────────────────────────────

@routes.route('/etp', methods=['GET'])
@require_auth()
def etp():
    mode       = request.args.get('mode', 'projet')  # 'projet' ou 'personne'
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    try:
        if mode == 'personne':
            # Filtre utilisateurs selon rôle
            if role == 'admin':
                user_where  = "u.actif = true"
                user_params = []
            else:
                # non-admin : uniquement soi-même
                user_where  = "u.actif = true AND u.id = %s"
                user_params = [user_id]

            rows = projet_service.db.fetch_all(
                "SELECT u.id, u.nom, u.prenom, u.email, u.role, "
                "s.nom as service_nom, s.code as service_code, "
                "COALESCE(s.is_unite, false) as is_unite, "
                "COUNT(t.id) as nb_taches, "
                "COALESCE(SUM(t.estimation_heures), 0) as heures_estimees, "
                "COALESCE(SUM(t.heures_reelles), 0) as heures_reelles "
                "FROM utilisateurs u "
                "LEFT JOIN services s ON s.id = u.service_id "
                "LEFT JOIN taches t ON t.assignee_id = u.id "
                "  AND t.statut NOT IN ('Terminé', 'Annulé') "
                f"WHERE {user_where} "
                "GROUP BY u.id, u.nom, u.prenom, u.email, u.role, "
                "  s.nom, s.code, s.is_unite "
                "ORDER BY heures_estimees DESC, u.nom",
                user_params
            )
            result = [dict(r) for r in rows] if rows else []
            HEURES_AN = 1540  # 1 ETP = ~220 jours * 7h
            for r in result:
                h = float(r.get('heures_estimees') or 0)
                r['heures_dispo'] = round(HEURES_AN - h, 1)
                r['pct_charge']   = round(h / HEURES_AN * 100, 1)
                r['etp_charge']   = round(h / HEURES_AN, 2)
            return jsonify({"list": result, "mode": "personne",
                            "heures_an": HEURES_AN})
        else:
            # Charge par projet — filtré par ownership
            if role == 'admin':
                proj_where  = "p.statut NOT IN ('Terminé', 'Annulé')"
                proj_params = []
            else:
                own_w, own_p = _ownership_where(user_id, role, service_id, 'p')
                proj_where   = f"p.statut NOT IN ('Terminé', 'Annulé') AND ({own_w})"
                proj_params  = own_p
            rows = projet_service.db.fetch_all(
                "SELECT p.id, p.code, p.nom, p.statut, "
                "COALESCE(SUM(t.estimation_heures), 0) as heures_estimees, "
                "COALESCE(SUM(t.heures_reelles), 0) as heures_reelles, "
                "COUNT(t.id) as nb_taches "
                "FROM projets p "
                "LEFT JOIN taches t ON t.projet_id = p.id "
                f"WHERE {proj_where} "
                "GROUP BY p.id, p.code, p.nom, p.statut "
                "ORDER BY heures_estimees DESC",
                proj_params
            )
            result = [dict(r) for r in rows] if rows else []
            total_h = sum(r.get('heures_estimees', 0) or 0 for r in result)
            return jsonify({
                "list": result, "mode": "projet",
                "total_heures": total_h,
                "total_jours": round(total_h / 7, 1),
                "total_etp": round(total_h / 154, 2),
            })
    except Exception as e:
        return jsonify({"list": [], "error": str(e)})


@routes.route('/users/actifs', methods=['GET'])
@require_auth()
def get_users_actifs():
    """Tous les utilisateurs actifs — pour les selects tâches/équipe."""
    try:
        rows = tache_service.db.fetch_all(
            "SELECT u.id, u.nom, u.prenom, u.email, u.role, "
            "u.service_id, s.nom as service_nom, s.code as service_code, "
            "COALESCE(s.is_unite, false) as is_unite "
            "FROM utilisateurs u "
            "LEFT JOIN services s ON s.id = u.service_id "
            "WHERE u.actif = true ORDER BY u.nom, u.prenom"
        )
        result = [dict(r) for r in rows] if rows else []
        return jsonify({"list": result})
    except Exception as e:
        return jsonify({"list": [], "error": str(e)})


# ─────────────────────────────────────────────
# NOTIFICATIONS
# ─────────────────────────────────────────────

@routes.route('/notifications', methods=['GET'])
@require_auth()
def get_notifications():
    try:
        rows = budget_service.db.fetch_all(
            "SELECT * FROM notifications ORDER BY date_creation DESC LIMIT 50"
        )
        result = [dict(r) for r in rows] if rows else []
        non_lues = sum(1 for r in result if not r.get('lue'))
        return jsonify({"list": result, "non_lues": non_lues})
    except Exception as e:
        return jsonify({"list": [], "non_lues": 0})

@routes.route('/notifications/<int:notif_id>/lire', methods=['POST'])
@require_auth()
def lire_notification(notif_id):
    try:
        budget_service.db.execute("UPDATE notifications SET lue=true WHERE id=%s", [notif_id])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400


@routes.route('/notifications/generate', methods=['POST'])
@require_auth()
def generate_notifications():
    """Génère automatiquement les notifications pour tâches en retard,
    contrats expirant bientôt et BC en attente trop longtemps."""
    db = budget_service.db
    created = 0
    today = __import__('datetime').date.today()

    def _upsert(ref_type, ref_id, titre, message, niveau):
        """Insère une notif uniquement si elle n'existe pas déjà (non lue)."""
        nonlocal created
        existing = db.fetch_one(
            "SELECT id FROM notifications WHERE ref_type=%s AND ref_id=%s AND lue=false",
            [ref_type, ref_id]
        )
        if not existing:
            db.execute(
                "INSERT INTO notifications (titre, message, lue, ref_type, ref_id, niveau) "
                "VALUES (%s, %s, false, %s, %s, %s)",
                [titre, message, ref_type, ref_id, niveau]
            )
            created += 1

    # ── Tâches en retard ──────────────────────────────────────────
    try:
        rows = db.fetch_all(
            "SELECT t.id, t.titre, t.date_echeance, p.nom as projet_nom "
            "FROM taches t LEFT JOIN projets p ON p.id = t.projet_id "
            "WHERE t.date_echeance < %s AND t.statut NOT IN ('Terminé','ANNULE')",
            [today]
        )
        for r in (rows or []):
            _upsert(
                'tache', r['id'],
                f"Tâche en retard : {r['titre']}",
                f"Échéance dépassée ({r['date_echeance']}) — Projet : {r['projet_nom'] or '—'}",
                'URGENT'
            )
    except Exception as _e:
        pass

    # ── Contrats expirant dans les 30 jours ───────────────────────
    try:
        rows = db.fetch_all(
            "SELECT id, objet, date_fin FROM contrats "
            "WHERE date_fin BETWEEN %s AND %s AND statut IN ('ACTIF','RECONDUIT')",
            [today, today + __import__('datetime').timedelta(days=30)]
        )
        for r in (rows or []):
            _upsert(
                'contrat', r['id'],
                f"Contrat expirant bientôt : {r['objet']}",
                f"Date de fin : {r['date_fin']}",
                'ALERTE'
            )
    except Exception as _e:
        pass

    # ── BC en attente depuis plus de 15 jours ─────────────────────
    try:
        rows = db.fetch_all(
            "SELECT id, objet, date_creation FROM bons_commande "
            "WHERE statut = 'EN_ATTENTE' AND date_creation < %s",
            [today - __import__('datetime').timedelta(days=15)]
        )
        for r in (rows or []):
            _upsert(
                'bc', r['id'],
                f"BC en attente depuis plus de 15 jours : {r['objet']}",
                f"Créé le {r['date_creation']} — en attente de validation",
                'ALERTE'
            )
    except Exception as _e:
        pass

    return jsonify({"success": True, "created": created})


# ─────────────────────────────────────────────
# GANTT
# ─────────────────────────────────────────────

@routes.route('/gantt', methods=['GET'])
@require_auth()
def get_gantt():
    """Retourne projets + tâches pour le diagramme de Gantt."""
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    svc_id     = g.user.get('service_id')

    service_id = request.args.get('service_id', type=int)
    date_debut = request.args.get('date_debut')   # YYYY-MM-DD
    date_fin   = request.args.get('date_fin')     # YYYY-MM-DD
    projet_id  = request.args.get('projet_id', type=int)

    db = budget_service.db

    # ── Projets ─────────────────────────────────────────────
    proj_where = ["(p.date_debut IS NOT NULL OR p.date_fin_prevue IS NOT NULL)"]
    proj_params = []
    # Filtre propriétaire (non-admin)
    if role != 'admin':
        own_w, own_p = _ownership_where(user_id, role, svc_id, 'p')
        proj_where.append(f"({own_w})")
        proj_params.extend(own_p)
    if service_id:
        proj_where.append("p.service_id = %s")
        proj_params.append(service_id)
    if date_debut:
        proj_where.append("(p.date_fin_prevue IS NULL OR p.date_fin_prevue >= %s)")
        proj_params.append(date_debut)
    if date_fin:
        proj_where.append("(p.date_debut IS NULL OR p.date_debut <= %s)")
        proj_params.append(date_fin)
    proj_sql = (
        "SELECT p.id, p.code, p.nom, p.date_debut, p.date_fin_prevue, "
        "p.avancement, p.statut, s.nom as service_nom "
        "FROM projets p "
        "LEFT JOIN services s ON s.id = p.service_id "
        "WHERE " + " AND ".join(proj_where) +
        " ORDER BY p.date_debut NULLS LAST"
    )
    try:
        projets = db.fetch_all(proj_sql, proj_params) or []
    except Exception:
        projets = []

    # ── Tâches ──────────────────────────────────────────────
    tache_where = ["t.date_echeance IS NOT NULL"]
    tache_params = []
    # Filtre visibilité tâches (non-admin)
    if role != 'admin':
        t_w, t_p = _tache_visibility_where(user_id, role, svc_id)
        tache_where.append(f"({t_w})")
        tache_params.extend(t_p)
    if projet_id:
        tache_where.append("t.projet_id = %s")
        tache_params.append(projet_id)
    elif service_id:
        tache_where.append("p.service_id = %s")
        tache_params.append(service_id)
    if date_debut:
        tache_where.append("(t.date_echeance IS NULL OR t.date_echeance >= %s)")
        tache_params.append(date_debut)
    if date_fin:
        tache_where.append("(t.date_debut IS NULL OR t.date_debut <= %s)")
        tache_params.append(date_fin)
    tache_sql = (
        "SELECT t.id, t.titre, t.statut, t.priorite, t.avancement, "
        "t.date_debut, t.date_echeance, t.responsable_label, "
        "t.projet_id, p.nom as projet_nom, p.code as projet_code "
        "FROM taches t "
        "JOIN projets p ON p.id = t.projet_id "
        "WHERE " + " AND ".join(tache_where) +
        " ORDER BY t.date_debut NULLS LAST, t.date_echeance"
    )
    try:
        taches = db.fetch_all(tache_sql, tache_params) or []
    except Exception:
        taches = []

    def _str(v):
        return str(v) if v is not None else None

    return jsonify({
        "projets": [
            {**dict(r), "date_debut": _str(r.get("date_debut")),
             "date_fin_prevue": _str(r.get("date_fin_prevue"))}
            for r in projets
        ],
        "taches": [
            {**dict(r), "date_debut": _str(r.get("date_debut")),
             "date_echeance": _str(r.get("date_echeance"))}
            for r in taches
        ],
    })


# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────

@routes.route('/export/budget', methods=['GET'])
@require_auth()
def export_budget():
    import io
    import traceback
    from datetime import datetime
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        return jsonify({"error": f"openpyxl non installé: {e}"}), 500
    from flask import send_file

    exercice  = request.args.get('exercice', datetime.now().year, type=int)
    next_year = exercice + 1
    db        = budget_service.db
    BLUE      = "2563A8"

    # ── helpers ────────────────────────────────────────────────
    def hdr(ws, row_no):
        for cell in ws[row_no]:
            if cell.value is not None:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor=BLUE)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)

    def auto_w(ws, mn=10, mx=50):
        for col in ws.columns:
            w = max(mn, min(mx, max((len(str(c.value or "")) for c in col), default=0) + 2))
            ws.column_dimensions[get_column_letter(col[0].column)].width = w

    def fdate(v):
        return v.strftime("%Y-%m-%d") if v and hasattr(v, "strftime") else (str(v)[:10] if v else None)

    def ff(v):
        try:    return float(v or 0)
        except: return 0.0

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Feuille 1 : Synthèse ────────────────────────────────────
        ws1 = wb.create_sheet(f"Synthèse {exercice}")
        ws1.append([f"SYNTHÈSE BUDGET DSI — Exercice {exercice}"])
        ws1["A1"].font = Font(bold=True, size=14, color=BLUE)
        ws1.merge_cells("A1:G1")
        ws1.append([f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws1.append([])
        ws1.append(["Entité", "Nature", "Prévisionnel", "Voté", "Engagé", "Solde", "Statut"])
        hdr(ws1, 4)

        rows = db.fetch_all(
            "SELECT e.nom as entite_nom, b.nature, "
            "COALESCE(b.montant_previsionnel,0) as montant_prevu, "
            "COALESCE(b.montant_vote,0) as montant_vote, "
            "COALESCE(b.montant_engage,0) as montant_engage, "
            "COALESCE(b.montant_solde,0) as montant_solde, b.statut "
            "FROM budgets_annuels b "
            "LEFT JOIN entites e ON e.id = b.entite_id "
            "WHERE b.exercice = %s ORDER BY e.nom, b.nature",
            [exercice]
        ) or []
        for r in rows:
            ws1.append([r["entite_nom"], r["nature"],
                        ff(r["montant_prevu"]), ff(r["montant_vote"]),
                        ff(r["montant_engage"]), ff(r["montant_solde"]),
                        r["statut"]])
        for row in ws1.iter_rows(min_row=5, min_col=3, max_col=6):
            for c in row:
                c.number_format = "#,##0.00"
        auto_w(ws1)

        # ── Feuille 2 : Lignes ──────────────────────────────────────
        ws2 = wb.create_sheet(f"Lignes {exercice}")
        lignes = db.fetch_all(
            "SELECT e.nom as entite_nom, b.nature, "
            "l.libelle, l.nature as ligne_nature, "
            "a.nom as application_nom, f.nom as fournisseur_nom, "
            "l.note as ref_dsi, "
            "COALESCE(l.montant_vote,0) as montant_vote, "
            "COALESCE(l.montant_engage,0) as montant_engage, "
            "COALESCE(l.montant_solde,0) as montant_solde, "
            "CASE WHEN COALESCE(l.montant_vote,0)>0 "
            "  THEN ROUND((COALESCE(l.montant_engage,0)/l.montant_vote*100)::numeric,1) "
            "  ELSE 0 END as taux_pct, "
            "CASE WHEN COALESCE(l.montant_vote,0)>0 "
            "      AND COALESCE(l.montant_engage,0) > l.montant_vote "
            "  THEN 'DEPASSE' "
            "  WHEN COALESCE(l.montant_vote,0)>0 "
            "       AND COALESCE(l.montant_engage,0) >= l.montant_vote*0.9 "
            "  THEN 'SEUIL' "
            "  ELSE 'OK' END as alerte "
            "FROM lignes_budgetaires l "
            "JOIN budgets_annuels b ON b.id = l.budget_id "
            "LEFT JOIN entites e ON e.id = b.entite_id "
            "LEFT JOIN applications a ON a.id = l.application_id "
            "LEFT JOIN fournisseurs f ON f.id = l.fournisseur_id "
            "WHERE b.exercice = %s ORDER BY e.nom, b.nature, l.libelle",
            [exercice]
        ) or []

        cur_grp = None
        for l in lignes:
            grp = f"{l['entite_nom'] or '-'}  -  {l['nature'] or '-'}  -  {exercice}"
            if grp != cur_grp:
                if cur_grp is not None:
                    ws2.append([])
                ws2.append([grp])
                ws2[f"A{ws2.max_row}"].font = Font(bold=True, size=11, color=BLUE)
                ws2.append(["Libelle", "Nature", "Application", "Fournisseur",
                            "Reference DSI", "Vote", "Engage", "Solde", "Taux %", "Alerte"])
                hdr(ws2, ws2.max_row)
                cur_grp = grp
            ws2.append([
                l["libelle"], l["ligne_nature"] or l["nature"],
                l["application_nom"], l["fournisseur_nom"],
                l["ref_dsi"],
                ff(l["montant_vote"]), ff(l["montant_engage"]),
                ff(l["montant_solde"]), ff(l["taux_pct"]) / 100,
                l["alerte"]
            ])

        for row in ws2.iter_rows(min_row=1):
            for c in row:
                if isinstance(c.value, float):
                    if c.column in (6, 7, 8):
                        c.number_format = "#,##0.00"
                    elif c.column == 9:
                        c.number_format = "0.0%"
        auto_w(ws2)

        # ── Feuille 3 : Contrats actifs ─────────────────────────────
        ws3 = wb.create_sheet("Contrats actifs")
        ws3.append(["Entite", "N Contrat", "Objet", "Fournisseur",
                    "Application", "Date fin", "Jours", "Montant HT",
                    "Montant max", "Engage", "Alerte"])
        hdr(ws3, 1)

        contrats = db.fetch_all(
            "SELECT "
            "(SELECT e.code FROM entites e "
            " JOIN bons_commande bc2 ON bc2.entite_id = e.id "
            " WHERE bc2.contrat_id = c.id LIMIT 1) as entite_code, "
            "c.numero_contrat, c.objet, f.nom as fournisseur_nom, "
            "(SELECT a.nom FROM applications a "
            " JOIN lignes_budgetaires lb ON lb.application_id = a.id "
            " JOIN bons_commande bc2 ON bc2.ligne_budgetaire_id = lb.id "
            " WHERE bc2.contrat_id = c.id LIMIT 1) as application_nom, "
            "c.date_fin, (c.date_fin::date - CURRENT_DATE) as jours, "
            "COALESCE(c.montant_initial_ht,0) as montant_ht, "
            "COALESCE(c.montant_total_ht,0) as montant_max, "
            "(SELECT COALESCE(SUM(bc2.montant_ttc),0) FROM bons_commande bc2 "
            " WHERE bc2.contrat_id = c.id) as montant_engage "
            "FROM contrats c "
            "LEFT JOIN fournisseurs f ON f.id = c.fournisseur_id "
            "WHERE c.statut IN ('ACTIF','RECONDUIT') ORDER BY c.date_fin ASC"
        ) or []

        for c in contrats:
            j = int(c["jours"]) if c["jours"] is not None else None
            if j is None:               alerte = "OK"
            elif j < 0:                 alerte = "EXPIRE"
            elif j <= 30:               alerte = "CRITIQUE"
            elif j <= 90:               alerte = "ATTENTION"
            elif j <= 180:              alerte = "INFO"
            else:                       alerte = "OK"
            ws3.append([c["entite_code"], c["numero_contrat"], c["objet"],
                        c["fournisseur_nom"], c["application_nom"],
                        fdate(c["date_fin"]), j,
                        ff(c["montant_ht"]), ff(c["montant_max"]),
                        ff(c["montant_engage"]), alerte])
        for row in ws3.iter_rows(min_row=2, min_col=9, max_col=11):
            for c in row:
                c.number_format = "#,##0.00"
        auto_w(ws3)

        # ── Feuille 4 : BC ──────────────────────────────────────────
        ws4 = wb.create_sheet(f"BC {exercice}")
        ws4.append(["Entite", "N BC", "Date", "Fournisseur", "Objet",
                    "Contrat", "Ligne budgetaire", "Application", "HT", "TTC", "Statut"])
        hdr(ws4, 1)

        bcs = db.fetch_all(
            "SELECT e.code as entite_code, bc.numero_bc, bc.date_creation, "
            "f.nom as fournisseur_nom, bc.objet, c.numero_contrat, "
            "lb.libelle as ligne_libelle, a.nom as application_nom, "
            "COALESCE(bc.montant_ht,0) as montant_ht, "
            "COALESCE(bc.montant_ttc,0) as montant_ttc, bc.statut "
            "FROM bons_commande bc "
            "LEFT JOIN entites e ON e.id = bc.entite_id "
            "LEFT JOIN fournisseurs f ON f.id = bc.fournisseur_id "
            "LEFT JOIN contrats c ON c.id = bc.contrat_id "
            "LEFT JOIN lignes_budgetaires lb ON lb.id = bc.ligne_budgetaire_id "
            "LEFT JOIN applications a ON a.id = lb.application_id "
            "WHERE EXTRACT(YEAR FROM bc.date_creation) = %s "
            "ORDER BY bc.date_creation DESC",
            [exercice]
        ) or []
        for bc in bcs:
            ws4.append([bc["entite_code"], bc["numero_bc"], fdate(bc["date_creation"]),
                        bc["fournisseur_nom"], bc["objet"], bc["numero_contrat"],
                        bc["ligne_libelle"], bc["application_nom"],
                        ff(bc["montant_ht"]), ff(bc["montant_ttc"]), bc["statut"]])
        for row in ws4.iter_rows(min_row=2, min_col=9, max_col=10):
            for c in row:
                c.number_format = "#,##0.00"
        auto_w(ws4)

        # ── Feuille 5 : Prévisionnel N+1 ────────────────────────────
        ws5 = wb.create_sheet(f"Previsionnel {next_year}")
        ws5.append([f"BUDGET PREVISIONNEL {next_year} - DSI"])
        ws5["A1"].font = Font(bold=True, size=14, color=BLUE)
        ws5.merge_cells("A1:H1")
        ws5.append([f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                    f"Source : Donnees reelles {next_year}"])
        ws5.append([])

        prev = db.fetch_all(
            "SELECT e.nom as entite_nom, b.nature, l.libelle, "
            "a.nom as application_nom, f.nom as fournisseur_nom, "
            "COALESCE(l.montant_prevu,0) as montant_prevu, l.note as ref_dsi "
            "FROM lignes_budgetaires l "
            "JOIN budgets_annuels b ON b.id = l.budget_id "
            "LEFT JOIN entites e ON e.id = b.entite_id "
            "LEFT JOIN applications a ON a.id = l.application_id "
            "LEFT JOIN fournisseurs f ON f.id = l.fournisseur_id "
            "WHERE b.exercice = %s ORDER BY e.nom, b.nature, l.libelle",
            [next_year]
        ) or []

        cur_grp = None
        grp_total = {}
        for l in prev:
            grp = f"{l['entite_nom'] or '-'}  -  {l['nature'] or '-'}  -  {next_year}"
            if grp != cur_grp:
                if cur_grp is not None:
                    ws5.append([f"SOUS-TOTAL  {cur_grp}", None, None, None, None,
                                 grp_total.get(cur_grp, 0)])
                    ws5.cell(ws5.max_row, 1).font = Font(bold=True)
                    ws5.cell(ws5.max_row, 6).number_format = "#,##0.00"
                    ws5.append([])
                ws5.append([grp])
                ws5[f"A{ws5.max_row}"].font = Font(bold=True, size=11, color=BLUE)
                ws5.append(["Entite", "Nature", "Libelle", "Application", "Fournisseur",
                            f"Montant prevu N+1", "Reference DSI", "Note"])
                hdr(ws5, ws5.max_row)
                cur_grp = grp
                grp_total[grp] = 0
            montant = ff(l["montant_prevu"])
            grp_total[grp] += montant
            ws5.append([l["entite_nom"], l["nature"], l["libelle"],
                        l["application_nom"], l["fournisseur_nom"],
                        montant, l["ref_dsi"], None])
            ws5.cell(ws5.max_row, 6).number_format = "#,##0.00"

        if cur_grp:
            ws5.append([f"SOUS-TOTAL  {cur_grp}", None, None, None, None,
                         grp_total.get(cur_grp, 0)])
            ws5.cell(ws5.max_row, 1).font = Font(bold=True)
            ws5.cell(ws5.max_row, 6).number_format = "#,##0.00"
            ws5.append([])
            ws5.append([f"TOTAL GENERAL BUDGET PREVISIONNEL {next_year}",
                        None, None, None, None, sum(grp_total.values())])
            ws5.cell(ws5.max_row, 1).font = Font(bold=True, size=12)
            ws5.cell(ws5.max_row, 6).number_format = "#,##0.00"
        auto_w(ws5)

        # ── Envoi ────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Budget_DSI_{exercice}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


# ─────────────────────────────────────────────
# MODULES (activation plugins — admin)
# ─────────────────────────────────────────────

@routes.route('/modules', methods=['GET'])
def get_modules_public():
    """Public : retourne les modules et leur statut (utilisé par le frontend au login)."""
    try:
        rows = auth_service.db.fetch_all(
            "SELECT module_name, enabled FROM modules_config ORDER BY module_name"
        )
        return jsonify({"list": rows or []})
    except Exception:
        return jsonify({"list": []})


@routes.route('/admin/modules', methods=['GET'])
@require_auth('admin')
def get_modules_admin():
    rows = auth_service.db.fetch_all(
        "SELECT module_name, enabled, date_activation FROM modules_config ORDER BY module_name"
    )
    return jsonify({"list": rows or []})


@routes.route('/admin/modules/<string:module_name>', methods=['PUT'])
@require_auth('admin')
def toggle_module(module_name):
    data = request.json or {}
    enabled = bool(data.get('enabled', False))
    try:
        auth_service.db.execute("""
            INSERT INTO modules_config (module_name, enabled, date_activation)
            VALUES (%s, %s, NOW())
            ON CONFLICT (module_name) DO UPDATE
                SET enabled=EXCLUDED.enabled, date_activation=NOW()
        """, [module_name, enabled])
        return _ok(enabled=enabled)
    except Exception as e:
        return _err(e)


# ── SMTP config & test ─────────────────────────────────────────────────────────

@routes.route('/admin/smtp/config', methods=['GET'])
@require_auth('admin')
def get_smtp_config():
    """Retourne la config SMTP (sans le mot de passe)."""
    import os
    return jsonify({
        'host':    os.getenv('SMTP_HOST', ''),
        'port':    os.getenv('SMTP_PORT', '587'),
        'user':    os.getenv('SMTP_USER', ''),
        'from':    os.getenv('SMTP_FROM', ''),
        'tls':     os.getenv('SMTP_TLS', 'true'),
        'configured': bool(os.getenv('SMTP_HOST') and os.getenv('SMTP_USER')),
    })


@routes.route('/admin/smtp/test', methods=['POST'])
@require_auth('admin')
def test_smtp():
    """Teste la connexion SMTP et envoie un mail de test."""
    import os, smtplib, ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    data = request.json or {}
    to_email = data.get('to_email', '').strip()
    if not to_email:
        return _err("Email destinataire requis")

    host   = os.getenv('SMTP_HOST', '')
    port   = int(os.getenv('SMTP_PORT', '587'))
    user   = os.getenv('SMTP_USER', '')
    passwd = os.getenv('SMTP_PASS', '')
    from_  = os.getenv('SMTP_FROM', user)
    use_tls = os.getenv('SMTP_TLS', 'true').lower() not in ('false', '0', 'no')

    if not host:
        return _err("SMTP_HOST non configuré dans les variables d'environnement")
    if not user:
        return _err("SMTP_USER non configuré dans les variables d'environnement")

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = '[Budget Manager] Test de connexion SMTP'
        msg['From']    = from_
        msg['To']      = to_email
        html = """<html><body>
            <p>✅ <strong>La connexion SMTP fonctionne correctement.</strong></p>
            <p>Ce message a été envoyé automatiquement par Budget Manager Pro pour valider la configuration.</p>
            <hr><small>Serveur : {host}:{port} — TLS : {tls}</small>
        </body></html>""".format(host=host, port=port, tls='oui' if use_tls else 'non')
        msg.attach(MIMEText(html, 'html', 'utf-8'))

        if use_tls:
            context = ssl.create_default_context()
            with smtplib.SMTP(host, port, timeout=10) as smtp:
                smtp.ehlo()
                smtp.starttls(context=context)
                smtp.ehlo()
                if passwd:
                    smtp.login(user, passwd)
                smtp.sendmail(from_, [to_email], msg.as_string())
        else:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=context, timeout=10) as smtp:
                if passwd:
                    smtp.login(user, passwd)
                smtp.sendmail(from_, [to_email], msg.as_string())

        return _ok(message=f"Email de test envoyé à {to_email}")

    except smtplib.SMTPAuthenticationError:
        return _err("Échec d'authentification — vérifiez SMTP_USER et SMTP_PASS")
    except smtplib.SMTPConnectError:
        return _err(f"Impossible de se connecter à {host}:{port}")
    except TimeoutError:
        return _err(f"Timeout — {host}:{port} ne répond pas (vérifiez le port et le pare-feu)")
    except Exception as e:
        return _err(f"Erreur SMTP : {e}")


# ─────────────────────────────────────────────
# NOTES
# ─────────────────────────────────────────────

_db_notes = None
def _notes_db():
    global _db_notes
    if _db_notes is None:
        from app.services.database_service import DatabaseService
        _db_notes = DatabaseService()
    return _db_notes

def _notes_ownership(user_id, role, service_id):
    if role == 'admin':
        return "1=1", []
    return "n.created_by_id = %s", [user_id]

@routes.route('/note', methods=['GET'])
@require_auth()
def get_notes():
    user_id    = g.user.get('sub')
    role       = g.user.get('role')
    service_id = g.user.get('service_id')
    db = _notes_db()
    where, params = _notes_ownership(user_id, role, service_id)
    note_type = request.args.get('type')
    projet_id = request.args.get('projet_id')
    if note_type:
        where += " AND n.type = %s"; params.append(note_type)
    if projet_id:
        where += " AND n.projet_id = %s"; params.append(int(projet_id))
    rows = db.fetch_all(
        "SELECT n.*, p.nom as projet_nom, p.code as projet_code, "
        "u.nom || ' ' || u.prenom as auteur_nom "
        "FROM notes n "
        "LEFT JOIN projets p ON p.id = n.projet_id "
        "LEFT JOIN utilisateurs u ON u.id = n.created_by_id "
        f"WHERE {where} ORDER BY n.updated_at DESC",
        params
    )
    return jsonify({"count": len(rows or []), "list": [dict(r) for r in (rows or [])]})

@routes.route('/note', methods=['POST'])
@require_auth()
def create_note():
    data    = request.json or {}
    user_id = g.user.get('sub')
    db = _notes_db()
    try:
        row = db.execute_returning(
            "INSERT INTO notes (titre, contenu, type, projet_id, couleur, created_by_id, updated_at) "
            "VALUES (%s, %s, %s, %s, %s, %s, NOW()) RETURNING id",
            [data.get('titre') or 'Sans titre',
             data.get('contenu') or '',
             data.get('type', 'postit'),
             data.get('projet_id') or None,
             data.get('couleur') or '#fff9c4',
             user_id]
        )
        return jsonify({"success": True, "id": row[0]}), 201
    except Exception as e:
        return _err(str(e))

@routes.route('/note/<int:note_id>', methods=['PUT'])
@require_auth()
def update_note(note_id):
    data    = request.json or {}
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    db = _notes_db()
    existing = db.fetch_one("SELECT created_by_id FROM notes WHERE id = %s", [note_id])
    if not existing:
        return _err("Note introuvable", 404)
    if role != 'admin' and existing['created_by_id'] != user_id:
        return _err("Accès refusé", 403)
    try:
        db.execute(
            "UPDATE notes SET titre=%s, contenu=%s, type=%s, projet_id=%s, couleur=%s, updated_at=NOW() WHERE id=%s",
            [data.get('titre') or 'Sans titre',
             data.get('contenu') or '',
             data.get('type', 'postit'),
             data.get('projet_id') or None,
             data.get('couleur') or '#fff9c4',
             note_id]
        )
        return _ok()
    except Exception as e:
        return _err(str(e))

@routes.route('/note/<int:note_id>', methods=['DELETE'])
@require_auth()
def delete_note(note_id):
    user_id = g.user.get('sub')
    role    = g.user.get('role')
    db = _notes_db()
    existing = db.fetch_one("SELECT created_by_id FROM notes WHERE id = %s", [note_id])
    if not existing:
        return _err("Note introuvable", 404)
    if role != 'admin' and existing['created_by_id'] != user_id:
        return _err("Accès refusé", 403)
    try:
        db.execute("DELETE FROM notes WHERE id = %s", [note_id])
        return _ok()
    except Exception as e:
        return _err(str(e))


# ─────────────────────────────────────────────
# SSO — OAuth2 / OpenID Connect
# ─────────────────────────────────────────────

def _sso_db():
    from app.services.database_service import DatabaseService
    return DatabaseService()

def _sso_get_config():
    return _sso_db().fetch_one("SELECT * FROM sso_config WHERE id = 1")

def _sso_sign_state(value):
    from app.services.auth_service import SECRET_KEY
    window = str(int(time.time()) // 300)
    return hmac.new(SECRET_KEY.encode(), f"{value}:{window}".encode(), hashlib.sha256).hexdigest()

def _sso_verify_state(value, sig):
    from app.services.auth_service import SECRET_KEY
    ts = int(time.time()) // 300
    for t in [ts, ts - 1]:
        expected = hmac.new(SECRET_KEY.encode(), f"{value}:{t}".encode(), hashlib.sha256).hexdigest()
        if hmac.compare_digest(expected, sig):
            return True
    return False

def _oidc_discover(issuer_url):
    import requests as _req
    r = _req.get(issuer_url.rstrip('/') + '/.well-known/openid-configuration', timeout=8)
    r.raise_for_status()
    return r.json()


@routes.route('/auth/sso/config_public', methods=['GET'])
def sso_config_public():
    """Retourne si le SSO est activé — utilisé par la page de login (sans auth)."""
    cfg = _sso_get_config()
    if cfg and cfg.get('enabled') and cfg.get('issuer_url') and cfg.get('client_id'):
        return jsonify({'enabled': True, 'provider_name': cfg.get('provider_name') or 'SSO'})
    return jsonify({'enabled': False})


@routes.route('/auth/sso/login', methods=['GET'])
def sso_login():
    """Lance le flux OIDC — retourne l'URL d'autorisation du provider."""
    cfg = _sso_get_config()
    if not cfg or not cfg.get('enabled'):
        return jsonify({"error": "SSO non configuré ou désactivé"}), 404
    try:
        disc = _oidc_discover(cfg['issuer_url'])
    except Exception as e:
        return jsonify({"error": f"Provider inaccessible : {e}"}), 502
    state_val = secrets.token_urlsafe(16)
    state = f"{state_val}.{_sso_sign_state(state_val)}"
    params = {
        'response_type': 'code',
        'client_id':     cfg['client_id'],
        'redirect_uri':  cfg['redirect_uri'],
        'scope':         cfg.get('scope') or 'openid email profile',
        'state':         state,
    }
    return jsonify({'redirect_url': disc['authorization_endpoint'] + '?' + urlencode(params)})


@routes.route('/auth/sso/callback', methods=['GET'])
def sso_callback():
    """Callback OIDC : échange le code, émet un JWT interne, redirige vers le frontend."""
    import requests as _req

    error = request.args.get('error')
    if error:
        return flask_redirect(f"/?sso_error={error}", 302)

    code  = request.args.get('code', '')
    state = request.args.get('state', '')
    if not code:
        return flask_redirect("/?sso_error=missing_code", 302)

    parts = state.rsplit('.', 1)
    if len(parts) != 2 or not _sso_verify_state(parts[0], parts[1]):
        return flask_redirect("/?sso_error=invalid_state", 302)

    cfg = _sso_get_config()
    if not cfg or not cfg.get('enabled'):
        return flask_redirect("/?sso_error=sso_disabled", 302)

    try:
        disc = _oidc_discover(cfg['issuer_url'])
    except Exception:
        return flask_redirect("/?sso_error=provider_unreachable", 302)

    # Échange code → tokens
    try:
        token_resp = _req.post(disc['token_endpoint'], data={
            'grant_type':    'authorization_code',
            'code':          code,
            'redirect_uri':  cfg['redirect_uri'],
            'client_id':     cfg['client_id'],
            'client_secret': cfg['client_secret'],
        }, timeout=10)
        token_resp.raise_for_status()
        tokens = token_resp.json()
    except Exception:
        return flask_redirect("/?sso_error=token_exchange_failed", 302)

    access_token = tokens.get('access_token')
    if not access_token:
        return flask_redirect("/?sso_error=no_access_token", 302)

    # UserInfo
    try:
        ui_resp = _req.get(disc['userinfo_endpoint'],
                           headers={'Authorization': f'Bearer {access_token}'}, timeout=8)
        ui_resp.raise_for_status()
        userinfo = ui_resp.json()
    except Exception:
        return flask_redirect("/?sso_error=userinfo_failed", 302)

    email = (userinfo.get('email') or '').lower().strip()
    if not email:
        return flask_redirect("/?sso_error=no_email", 302)

    db = _sso_db()
    user = db.fetch_one(
        "SELECT id, nom, prenom, email, login, role, service_id, actif, modules "
        "FROM utilisateurs WHERE LOWER(email) = %s AND actif = true",
        [email]
    )

    if not user:
        if cfg.get('auto_create_users'):
            login = email.split('@')[0]
            if db.fetch_one("SELECT id FROM utilisateurs WHERE login=%s", [login]):
                login = email.replace('@', '_at_').replace('.', '_')
            role = cfg.get('default_role') or 'lecteur'
            from app.services.auth_service import AuthService
            mods = json.dumps(AuthService._DEFAULT_MODULES.get(role, []))
            db.execute(
                "INSERT INTO utilisateurs (nom, prenom, email, login, mot_de_passe, role, actif, modules) "
                "VALUES (%s, %s, %s, %s, '', %s, true, %s)",
                [userinfo.get('family_name', ''), userinfo.get('given_name', ''),
                 email, login, role, mods]
            )
            user = db.fetch_one(
                "SELECT id, nom, prenom, email, login, role, service_id, actif, modules "
                "FROM utilisateurs WHERE login=%s", [login]
            )
        else:
            return flask_redirect("/?sso_error=user_not_found", 302)

    # Émettre le JWT interne
    from app.services.auth_service import AuthService, SECRET_KEY, EXPIRY_HOURS
    import jwt as _jwt
    from datetime import datetime, timezone, timedelta
    modules = user.get('modules') or AuthService._DEFAULT_MODULES.get(user['role'], [])
    payload = {
        'sub':        str(user['id']),
        'login':      user['login'],
        'nom':        user.get('nom', ''),
        'prenom':     user.get('prenom', ''),
        'role':       user['role'],
        'service_id': user.get('service_id'),
        'modules':    modules,
        'iat':        datetime.now(timezone.utc),
        'exp':        datetime.now(timezone.utc) + timedelta(hours=EXPIRY_HOURS),
    }
    token = _jwt.encode(payload, SECRET_KEY, algorithm='HS256')
    return flask_redirect(f"/?sso_token={token}", 302)


@routes.route('/admin/sso/config', methods=['GET'])
@require_auth('admin')
def get_sso_config():
    cfg = _sso_get_config()
    if not cfg:
        return jsonify({'enabled': False, 'provider_name': 'SSO', 'issuer_url': '',
                        'client_id': '', 'client_secret': '', 'redirect_uri': '',
                        'scope': 'openid email profile', 'auto_create_users': False,
                        'default_role': 'lecteur'})
    result = dict(cfg)
    result['client_secret'] = '••••••••' if cfg.get('client_secret') else ''
    return jsonify(result)


@routes.route('/admin/sso/config', methods=['PUT'])
@require_auth('admin')
def save_sso_config():
    data = request.json or {}
    secret = data.get('client_secret', '')
    if secret in ('••••••••', ''):
        existing = _sso_get_config()
        secret = (existing.get('client_secret') or '') if existing else ''
    _sso_db().execute("""
        INSERT INTO sso_config
            (id, enabled, provider_name, issuer_url, client_id, client_secret,
             redirect_uri, scope, auto_create_users, default_role, date_maj)
        VALUES (1, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON CONFLICT (id) DO UPDATE SET
            enabled=%s, provider_name=%s, issuer_url=%s, client_id=%s, client_secret=%s,
            redirect_uri=%s, scope=%s, auto_create_users=%s, default_role=%s, date_maj=NOW()
    """, [
        data.get('enabled', False), data.get('provider_name', 'SSO'),
        data.get('issuer_url', ''), data.get('client_id', ''), secret,
        data.get('redirect_uri', ''), data.get('scope', 'openid email profile'),
        data.get('auto_create_users', False), data.get('default_role', 'lecteur'),
        data.get('enabled', False), data.get('provider_name', 'SSO'),
        data.get('issuer_url', ''), data.get('client_id', ''), secret,
        data.get('redirect_uri', ''), data.get('scope', 'openid email profile'),
        data.get('auto_create_users', False), data.get('default_role', 'lecteur'),
    ])
    return _ok()


@routes.route('/admin/sso/test', methods=['POST'])
@require_auth('admin')
def test_sso_discovery():
    """Teste la découverte OIDC depuis l'issuer_url fourni."""
    data = request.json or {}
    issuer = (data.get('issuer_url') or '').rstrip('/')
    if not issuer:
        return _err("issuer_url requis")
    try:
        disc = _oidc_discover(issuer)
        return jsonify({
            'success':                True,
            'authorization_endpoint': disc.get('authorization_endpoint'),
            'token_endpoint':         disc.get('token_endpoint'),
            'userinfo_endpoint':      disc.get('userinfo_endpoint'),
            'issuer':                 disc.get('issuer'),
        })
    except Exception as e:
        return _err(f"Échec découverte OIDC : {e}", 502)
